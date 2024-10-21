import streamlit as st

import pandas as pd
pd.options.display.max_columns=200

import numpy as np
import os, sys, glob
import humanize
import re
import regex
# import xlrd

import json
import itertools
#from urllib.request import urlopen
import requests
#import requests, xmltodict
import time, datetime
import math
from pprint import pprint
import gc
from tqdm import tqdm
tqdm.pandas()
import pickle

# import duckdb
# # https://stackoverflow.com/questions/75352219/fix-unimplemented-casting-error-in-duckdb-insert
# duckdb.default_connection.execute("SET GLOBAL pandas_analyze_sample=100000")
# import pyarrow

import logging
import zipfile
# import tarfile
# import py7zr
# import argparse

import warnings
warnings.filterwarnings("ignore")

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import units
from openpyxl.styles import Border, Side, PatternFill, GradientFill, Alignment
from openpyxl import drawing

# import matplotlib.pyplot as plt
# import seaborn as sns
# %matplotlib inline
# from matplotlib.colors import ListedColormap, BoundaryNorm

class Logger():
    def __init__(self, name = 'main',
                 strfmt = '[%(asctime)s] [%(levelname)s] > %(message)s', # strfmt = '[%(asctime)s] [%(name)s] [%(levelname)s] > %(message)s'
                 level = logging.INFO,
                 datefmt = '%H:%M:%S', # '%Y-%m-%d %H:%M:%S'
                #  datefmt = '%H:%M:%S %p %Z',

                 ):
        self.name = name
        self.strfmt = strfmt
        self.level = level
        self.datefmt = datefmt
        self.logger = logging.getLogger(name)
        self.logger.setLevel(self.level) #logging.INFO)
        self.offset = datetime.timezone(datetime.timedelta(hours=3))
        # create console handler and set level to debug
        self.ch = logging.StreamHandler()
        self.ch.setLevel(self.level)
        # create formatter
        self.strfmt = strfmt # '[%(asctime)s] [%(levelname)s] > %(message)s'
        self.datefmt = datefmt # '%H:%M:%S'
        # СЃРѕР·РґР°РµРј С„РѕСЂРјР°С‚С‚РµСЂ
        self.formatter = logging.Formatter(fmt=strfmt, datefmt=datefmt)
        self.formatter.converter = lambda *args: datetime.datetime.now(self.offset).timetuple()
        self.ch.setFormatter(self.formatter)
        # add ch to logger
        self.logger.addHandler(self.ch)
logger = Logger().logger
logger.propagate = False

if len(logger.handlers) > 1:
    for handler in logger.handlers:
        logger.removeHandler(handler)
    # del logger
    logger = Logger().logger
    logger.propagate = False

from utils_io import insert_pd_col_after_col, save_df_lst_to_excel_xlsxwriter, split_column_to_rows, clean_str


def read_okpd_dict(
    supp_dict_dir = '/content/cllct_rm_chars/data',
    fn = '20240624_ОКПД2_2024_09_13_1519.xlsx',
    sh_n = 'ОКПД2',
):
    try:
        okpd2_df = pd.read_excel(os.path.join(supp_dict_dir, fn), sheet_name=sh_n)
        logger.info(f"Справочник ОКПД2: (строк, колонок): {str(okpd2_df.shape)}")
        # display(okpd2_df.head(2))
    except Exception as err:
        logger.error(str(err))
    return okpd2_df

@st.cache_data(persist="disk")
def read_okpd_dict_fr_link(
        link = 'https://github.com/A-Gorby/cllct_rm_chars_st/raw/refs/heads/main/data/20240624_ОКПД2_2024_09_13_1519.xlsx',
    fn = '20240624_ОКПД2_2024_09_13_1519.xlsx',
    sh_n = 'ОКПД2',
):
    try:
        r = requests.get(link)
        # open('temp.xls', 'wb').write(r.content)
        open(fn, 'wb').write(r.content)
        
        # okpd2_df = pd.read_excel(link, sheet_name=sh_n, engine='openpyxl')
        # okpd2_df = pd.read_excel('temp.xls', sheet_name=sh_n, engine='openpyxl')
        okpd2_df = pd.read_excel(fn, sheet_name=sh_n, engine='openpyxl')
        # okpd2_df = pd.read_excel(link, sheet_name=sh_n)
        # logger.info(f"Справочник ОКПД2: (строк, колонок): {str(okpd2_df.shape)}")
        st.write(f"Справочник ОКПД2: (строк, колонок): {str(okpd2_df.shape)}")
        # display(okpd2_df.head(2))
    except Exception as err:
        logger.error(str(err))
        okpd2_df = pd.DataFrame()
    return okpd2_df
from openpyxl import load_workbook
from openpyxl.worksheet import merge
from openpyxl import utils

# def split_merged_cells(fn_path, sh_n_spgz, save_dir, debug=False):
#     wb = load_workbook(fn_path, read_only=False)
#     min_row = 3
#     # try:
#     ws = wb[sh_n_spgz]

#     if debug:
#         print(ws.merged_cells.ranges)
#     for merged_cells_range in  sorted(list(ws.merged_cells.ranges)):
#         if merged_cells_range.min_row <= min_row: continue
#         value = ws.cell(row=merged_cells_range.min_row, column=merged_cells_range.min_col).value
#         if debug:
#             print(merged_cells_range, value)
#         # ws.unmerge_cells(range_string = merged_cells_range)
#         ws.unmerge_cells(start_row=merged_cells_range.min_row, start_column=merged_cells_range.min_col, end_row=merged_cells_range.max_row, end_column=merged_cells_range.max_col)

#         for i_row in range(merged_cells_range.min_row, merged_cells_range.max_row + 1):
#             for i_col in range(merged_cells_range.min_col, merged_cells_range.max_col + 1):
#                 # ws.cell(row=i_row, column=i_col) = value
#                 # print(i_col, utils.cell.get_column_letter(i_col), i_row)
#                 # print(f"{utils.cell.get_column_letter(i_col)}{i_row}") #, ws[f"{l_col}{i_row}"])
#                 ws[f"{utils.cell.get_column_letter(i_col)}{i_row}"] = value
#     # except Exception as err:
#     #     print(err)

#     fn_proc_save = os.path.join(save_dir, fn_path.split(os.path.sep)[-1])
#     wb.save(fn_proc_save)
#     return fn_proc_save

def split_merged_cells(fn_path, sh_n_spgz, save_dir, debug=False):
    if not os.path.exists(fn_path):
        logger.error(f"Файл '{fn_path.split(os.path.sep)[-1]}' не найден")
        logger.error(f"Работа программы завершена")
        sys.exit(2)
    wb = load_workbook(fn_path, read_only=False)
    min_row = 3
    try:
        ws = wb[sh_n_spgz]

        if debug:
            print(ws.merged_cells.ranges)
        for merged_cells_range in  sorted(list(ws.merged_cells.ranges)):
            if merged_cells_range.min_row <= min_row: continue
            value = ws.cell(row=merged_cells_range.min_row, column=merged_cells_range.min_col).value
            if debug:
                print(merged_cells_range, value)
            # ws.unmerge_cells(range_string = merged_cells_range)
            ws.unmerge_cells(start_row=merged_cells_range.min_row, start_column=merged_cells_range.min_col, end_row=merged_cells_range.max_row, end_column=merged_cells_range.max_col)

            for i_row in range(merged_cells_range.min_row, merged_cells_range.max_row + 1):
                for i_col in range(merged_cells_range.min_col, merged_cells_range.max_col + 1):
                    # ws.cell(row=i_row, column=i_col) = value
                    # print(i_col, utils.cell.get_column_letter(i_col), i_row)
                    # print(f"{utils.cell.get_column_letter(i_col)}{i_row}") #, ws[f"{l_col}{i_row}"])
                    ws[f"{utils.cell.get_column_letter(i_col)}{i_row}"] = value
    except Exception as err:
        logger.error(str(err))
        logger.error(f"Работа программы завершена")
        sys.exit(2)

    fn_proc_save = os.path.join(save_dir, fn_path.split(os.path.sep)[-1])
    wb.save(fn_proc_save)
    return fn_proc_save
# fn_proc_save = split_merged_cells(fn_path, sh_n_spgz, save_dir=save_dir, debug=False)

def split_merged_cells_st(file_source, sh_n_spgz, save_suffix='_spliited', debug=False):
    
    wb = load_workbook(file_source, read_only=False)
    min_row = 3
    try:
        ws = wb[sh_n_spgz]

        if debug:
            st.write(ws.merged_cells.ranges)
        for merged_cells_range in  sorted(list(ws.merged_cells.ranges)):
            if merged_cells_range.min_row <= min_row: continue
            value = ws.cell(row=merged_cells_range.min_row, column=merged_cells_range.min_col).value
            if debug:
                st.write(merged_cells_range, value)
            # ws.unmerge_cells(range_string = merged_cells_range)
            ws.unmerge_cells(start_row=merged_cells_range.min_row, start_column=merged_cells_range.min_col, end_row=merged_cells_range.max_row, end_column=merged_cells_range.max_col)

            for i_row in range(merged_cells_range.min_row, merged_cells_range.max_row + 1):
                for i_col in range(merged_cells_range.min_col, merged_cells_range.max_col + 1):
                    # ws.cell(row=i_row, column=i_col) = value
                    # print(i_col, utils.cell.get_column_letter(i_col), i_row)
                    # print(f"{utils.cell.get_column_letter(i_col)}{i_row}") #, ws[f"{l_col}{i_row}"])
                    ws[f"{utils.cell.get_column_letter(i_col)}{i_row}"] = value
    except Exception as err:
        logger.error(str(err))
        logger.error(f"Работа программы завершена")
        sys.exit(2)

    fn_proc_save = os.path.basename(file_source.name).split('.xlsx')[0] + save_suffix + '.xlsx'
    if debug: st.write(fn_proc_save)
    wb.save(fn_proc_save)
    return fn_proc_save
# fn_proc_save = split_merged_cells_st(file_source, sh_n_spgz, save_suffix='_spliited', debug=False)

def split_merged_cells_in_dir(data_source_dir, save_dir, debug=False):
    fn_lst = glob.glob(os.path.join(data_source_dir,'*.xlsx'))
    print(fn_lst)
    # sh_n_kpgz = 'КПГЗ'
    sh_n_spgz = 'СПГЗ'
    for fn_path in fn_lst:
        print(fn_path)
        fn_proc_save = split_merged_cells(fn_path, sh_n_spgz, save_dir=save_dir, debug=False)
        logger.info(f"Файл '{fn_proc_save.split(os.path.sep)[-1]}' сохранен в папке '{save_dir}'")

def read_data(
    data_source_dir,
    fn_source,
    sh_n_source,
    debug=False
    ):


    if fn_source is None or sh_n_source is None:
        logger.info("Не опеределен входной файл/лист Excel")
        logger.info(f"Файл Excel: '{fn_source}'")
        logger.info(f"Лист Excel: '{sh_n_source}'")
    if not os.path.exists(os.path.join(data_source_dir, fn_source)):
        logger.error(f"Файл Excel: '{fn_source}' не найден")

    try:
        df_rm_source_t = pd.read_excel(os.path.join(data_source_dir, fn_source), sheet_name=sh_n_source, nrows=5, header=1)
        source_cols = list(df_rm_source_t.columns)
        converters = dict(zip(source_cols, len(source_cols)*[str]))
        df_rm_source = pd.read_excel(os.path.join(data_source_dir, fn_source), sheet_name=sh_n_source, header=1,
                                 converters=converters)
        logger.info(f"Файл Excel для обработки: '{fn_source}':\n(строк, колонок): {str(df_rm_source.shape)}")
    except Exception as err:
       logger.error(str(err))
       logger.error("Работа программы прекращена")
       sys.exit()

    return df_rm_source

def trim_right_dot_compress_spaces(s):
    if type(s)!=str: return s
    s = re.sub(r'\s{2,}', ' ', s.strip())
    if (len(s) > 0) and s[-1]=='.': s = s[:-1]
    return s

def delete_empty_rows_in_cell(s):
    """
    v 01.01 08.05.2024
    """
    if type(s)==str:
        ss = [v.strip() for v in s.split('\n')]
        ss = [v for v in ss if len(v) > 0]
        ss = '\n'.join(ss)
        return ss
    else:
        return s

def extract_kpgz_df_lst(fn, sh_n_kpgz, debug=False):
    kpgz_header_name_loc_df = pd.read_excel(fn, sheet_name=sh_n_kpgz, header=None, nrows=1 )
    # display(kpgz_header_name_loc_df)
    # print(kpgz_header_name_loc_df.values[0,0])
    kpgz_code_name = None
    try:
        kpgz_code_name_01 = re.sub( 'Позиция КПГЗ ', '', kpgz_header_name_loc_df.values[0,0]).strip()
    except Exception as err:
        print("ERROR: extract_kpgz_df_lst:")
        print(err)
    kpgz_code_name = kpgz_code_name_01

    kpgz_header_content_loc_df = pd.read_excel(fn, sheet_name=sh_n_kpgz, header=None, skiprows=1, nrows=7 )

    kpgz_header_content_loc_df.rename(columns = {0: 'Показатель', 1: 'Значение'}, inplace=True)

    kpgz_header_content_loc_df['КГПЗ Код Наименование'] = kpgz_code_name
    kpgz_header_content_loc_df = kpgz_header_content_loc_df[['КГПЗ Код Наименование'] + list(kpgz_header_content_loc_df.columns[:-1])]
    # display(kpgz_header_content_loc_df)

    kpgz_characteristics_name_loc_df = pd.read_excel(fn, sheet_name=sh_n_kpgz, header=None, skiprows=9, nrows=1)
        # display(kpgz_characteristics_name_loc_df)
    # print(kpgz_characteristics_name_loc_df.values[0,0])
    try:
        kpgz_code_name_02 = re.sub( 'Справочник характеристик и их значений позиции КПГЗ ', '', kpgz_characteristics_name_loc_df.values[0,0]).strip()
    except Exception as err:
        print(err)
    if kpgz_code_name_01 != kpgz_code_name_02:
        print(f"Наименование КПГЗ для заголовка '{kpgz_code_name_01}' не совпадает с наименованием для характеристик '{kpgz_code_name_02}'")


    kpgz_characteristics_content_loc_df = pd.read_excel(fn, sheet_name=sh_n_kpgz, skiprows=10) #, header=11)
    kpgz_characteristics_content_loc_df = kpgz_characteristics_content_loc_df[1:]
    kpgz_characteristics_content_loc_df['КГПЗ Код Наименование'] = kpgz_code_name
    # kpgz_characteristics_content_loc_df.columns = ['КГПЗ Код Наименование'] + list(kpgz_characteristics_content_loc_df.columns[:-1])
    kpgz_characteristics_content_loc_df = kpgz_characteristics_content_loc_df[['КГПЗ Код Наименование'] + list(kpgz_characteristics_content_loc_df.columns[:-1])]
    # display(kpgz_characteristics_content_loc_df.head())

    try:
        kpgz_characteristics_content_loc_df['Наименование характеристики'] = kpgz_characteristics_content_loc_df['Наименование характеристики'].apply(trim_right_dot_compress_spaces)
    except Exception as err:
        print(err)

    return kpgz_code_name, kpgz_header_content_loc_df, kpgz_characteristics_content_loc_df

def extract_spgz_df_lst_st(fn, sh_n_spgz, groupby_col='№п/п',
                        unique_test_cols=['Наименование СПГЗ', 'Единица измерения', 'ОКПД 2', 
                                          'Позиция КТРУ'
                                          # 'Код характеристики КТРУ'
                                          ],
                        significant_cols = ['Наименование характеристики', 'Единица измерения характеристики', 'Значение характеристики', 
                                            'Тип характеристики', 
                                            'Условная операция',
                                            'Тип выбора значений характеристики заказчиком',
                                            # 'Код характеристики КТРУ'
                                            'Код КТРУ характеристики',
                                            ],
                        debug=False,
                        ):

    """
    v01.02 08.05.2024
      изменения:
      1. после удаления зачеркнутого текста убрать пустые строки
      2. если вся ячейки по олокнам ниже в строке удалены то не добавлять и строку
      ['Наименование характеристики', 'Единица измерения характеристики', 'Значение характеристики', 'Тип характеристики',
      'Тип выбора значений характеристики заказчиком']
      3. отбор только тех строк, где 'Наименование СПГЗ' и  'Наименование характеристики' не пусты
      4. проверка правильности заполнения на листе СПГЗ колонок '№п/п', 'Наименование СПГЗ', 'Единица измерения'

    v01.03 21.10.2024
    изменения:
    'Код КТРУ характеристики'- он же 'Код характеристики КТРУ'

    """
    spgz_header_name_loc_df = pd.read_excel(fn, sheet_name=sh_n_spgz, header=None, nrows=1 )
    # display(kpgz_header_name_loc_df)
    # print(spgz_header_name_loc_df.values[0,0])
    spgz_code_name = None
    try:
        spgz_code_name = re.sub( 'Перечень позиций СПГЗ, относящихся к позиции КПГЗ ', '', spgz_header_name_loc_df.values[0,0]).strip()
    except Exception as err:
        print("ERROR: extract_spgz_df_lst:")
        print(err)

    spgz_characteristics_content_loc_df = pd.read_excel(fn, sheet_name=sh_n_spgz, header=1, #, skiprows=1
    converters = {'Наименование характеристики':str, 'Единица измерения характеристики':str, 'Значение характеристики':str, 'Тип характеристики':str,
      'Условная операция':str,                  
      'Тип выбора значений характеристики заказчиком':str,
      'Код КТРУ характеристики':str,
      'Позиция КТРУ':str,
                  }
                                                        )
    spgz_characteristics_content_loc_df = spgz_characteristics_content_loc_df[1:]
    # display(spgz_characteristics_content_loc_df.head())
    # spgz_characteristics_content_loc_df = spgz_characteristics_content_loc_df[
    #     spgz_characteristics_content_loc_df['Наименование СПГЗ'].notnull() & (spgz_characteristics_content_loc_df['Наименование СПГЗ'].str.len()>0) &
    #     spgz_characteristics_content_loc_df['Наименование характеристики'].notnull() & (spgz_characteristics_content_loc_df['Наименование характеристики'].str.len()>0)
    #     ]

    unnamed_cols = [col for col in spgz_characteristics_content_loc_df.columns if col.startswith('Unnamed:')]
    spgz_characteristics_content_loc_df.drop(columns=unnamed_cols, inplace=True)

    # проверка корректности заполнения по колонкам '№п/п', 'Наименование СПГЗ', 'Единица измерения'
    # количество непустых строк по ним должно совпадать поскольку они объединены это не всегда видно визуально

    npp_nunique = spgz_characteristics_content_loc_df['№п/п'].nunique()
    mask_for_value_counts = (spgz_characteristics_content_loc_df['№п/п'].notnull() |
    spgz_characteristics_content_loc_df['Наименование СПГЗ'].notnull() |
    spgz_characteristics_content_loc_df['Единица измерения'].notnull()
    )
    # что0нибудь непустое должно быть, если все пустые колонки - не считаем
    need_value_counts = spgz_characteristics_content_loc_df[mask_for_value_counts].value_counts(['№п/п', 'Наименование СПГЗ', 'Единица измерения'], dropna=False).shape[0]
    # name_spgz_nunique = spgz_characteristics_content_loc_df['Наименование СПГЗ'].nunique()
    # ei_nunique = spgz_characteristics_content_loc_df['Единица измерения'].nunique()
    # assert ((npp_nunique!=name_spgz_nunique) and (name_spgz_nunique != ei_nunique) and  (npp_nunique!=ei_nunique),
    #         "Ошибка заполнения объединенных ячеек по колонкам '№п/п', 'Наименование СПГЗ', 'Единица измерения'"
    #         )
    if debug: print(f"npp_nunique: '{npp_nunique}', need_value_counts: '{need_value_counts}'")
    if npp_nunique!= need_value_counts:
        print("Ошибка заполнения объединенных ячеек по колонкам '№п/п', 'Наименование СПГЗ', 'Единица измерения'")
        print(spgz_characteristics_content_loc_df[mask_for_value_counts].value_counts(['№п/п', 'Наименование СПГЗ', 'Единица измерения'], dropna=False))

    spgz_characteristics_content_loc_df['CГПЗ Код Наименование'] = spgz_code_name
    # spgz_characteristics_content_loc_df.columns = ['CГПЗ Код Наименование'] + list(spgz_characteristics_content_loc_df.columns[:-1])
    spgz_characteristics_content_loc_df = spgz_characteristics_content_loc_df[['CГПЗ Код Наименование'] + list(spgz_characteristics_content_loc_df.columns[:-1])]

    for col in unique_test_cols:
        spgz_characteristics_content_loc_df['Ошибка объединения ячеек\n' + col] = None

    spgz_characteristics_content_loc_df_groupped = spgz_characteristics_content_loc_df.groupby(groupby_col)

    for group_name, group_df in spgz_characteristics_content_loc_df_groupped:
        # print(group_name)
        # # проверка корректности заполнения по колонкам '№п/п', 'Наименование СПГЗ', 'Единица измерения'
        # # количество непустых строк по ним должно совпадать поскольку они объединены это не всегда видно визуально
        # npp_len = group_df.shape[0]
        # name_spgz_len = group_df['Наименование СПГЗ'].nunique()
        # ei_nunique = group_df['Единица измерения'].nunique()
        # assert ((npp_nunique!=name_spgz_nunique) and (name_spgz_nunique != ei_nunique) and  (npp_nunique!=ei_nunique),
        #         "Ошибка заполнения объединенных ячеек по колонкам '№п/п', 'Наименование СПГЗ', 'Единица измерения'"
        #         )
        # print(f"npp_nunique: '{npp_nunique}', name_spgz_nunique: '{name_spgz_nunique}', ei_nunique: '{ei_nunique}'")
        for col in unique_test_cols:
            # print(group_df[col].unique())
            if group_df[col].nunique()>1:
                group_err_str = str(group_df[col].unique().tolist())
                print(f"'{groupby_col}': '{group_name}'", f"Ошибка группировки (объединения) - в колонке '{col}'")
                print(group_err_str)
                # display(spgz_characteristics_content_loc_df[(spgz_characteristics_content_loc_df[groupby_col]==group_name)])
                spgz_characteristics_content_loc_df.loc[
                    (spgz_characteristics_content_loc_df[groupby_col]==group_name), 'Ошибка объединения ячеек\n' + col] = group_err_str
    # display(spgz_characteristics_content_loc_df.head())
    try:
        spgz_characteristics_content_loc_df['Наименование характеристики'] = spgz_characteristics_content_loc_df[
            'Наименование характеристики'].apply(trim_right_dot_compress_spaces)
    except Exception as err:
        print("ERROR: extract_spgz_df_lst: apply(trim_right_dot_compress_spaces)")
        print(err)
    # for col in significant_cols:
    for col in list(set(unique_test_cols + significant_cols)):
        try:
            spgz_characteristics_content_loc_df[col] = spgz_characteristics_content_loc_df[col].apply(
                delete_empty_rows_in_cell)
        except Exception as err:
            print(f"ERROR: extract_spgz_df_lst: apply(delete_empty_rows_in_cell): col: '{col}'")
            print(err)
    mask = (spgz_characteristics_content_loc_df['Наименование характеристики'].notnull() &
      (spgz_characteristics_content_loc_df['Наименование характеристики'].str.len()>0)
    )
    # mask = spgz_characteristics_content_loc_df[significant_cols[0]].notnull() & (spgz_characteristics_content_loc_df[significant_cols[0]].str.len()>0)
    # for col in significant_cols[1:]:
    #     mask = mask & spgz_characteristics_content_loc_df[col].notnull() & (spgz_characteristics_content_loc_df[col].str.len()>0)
    spgz_characteristics_content_loc_df = spgz_characteristics_content_loc_df[mask]

    return spgz_code_name, spgz_characteristics_content_loc_df

def extract_spgz_df_lst_st_v01_02(
        fn, sh_n_spgz, 
                        groupby_col='№п/п',
                        unique_test_cols=['Наименование СПГЗ', 'Единица измерения', 'ОКПД 2', 'Позиция КТРУ'],
                        significant_cols = ['Наименование характеристики', 'Единица измерения характеристики', 'Значение характеристики', 'Тип характеристики', 'Тип выбора значений характеристики заказчиком'],
                        debug=False,
                        ):

    """
    v01.02 08.05.2024
      изменения:
      1. после удаления зачеркнутого текста убрать пустые строки
      2. если вся ячейки по олокнам ниже в строке удалены то не добавлять и строку
      ['Наименование характеристики', 'Единица измерения характеристики', 'Значение характеристики', 'Тип характеристики',
      'Тип выбора значений характеристики заказчиком']
      3. отбор только тех строк, где 'Наименование СПГЗ' и  'Наименование характеристики' не пусты
      4. првоерка правильности заполнения на листе СПГЗ колонок '№п/п', 'Наименование СПГЗ', 'Единица измерения'

    """
    spgz_header_name_loc_df = pd.read_excel(fn, sheet_name=sh_n_spgz, header=None, nrows=1 )
    # display(kpgz_header_name_loc_df)
    # print(spgz_header_name_loc_df.values[0,0])
    spgz_code_name = None
    try:
        spgz_code_name = re.sub( 'Перечень позиций СПГЗ, относящихся к позиции КПГЗ ', '', spgz_header_name_loc_df.values[0,0]).strip()
    except Exception as err:
        print("ERROR: extract_spgz_df_lst:")
        print(err)

    spgz_characteristics_content_loc_df = pd.read_excel(fn, sheet_name=sh_n_spgz, header=1, #, skiprows=1
    converters = {'Наименование характеристики':str, 'Единица измерения характеристики':str, 'Значение характеристики':str, 'Тип характеристики':str,
      'Тип выбора значений характеристики заказчиком':str}
                                                        )
    spgz_characteristics_content_loc_df = spgz_characteristics_content_loc_df[1:]
    # display(spgz_characteristics_content_loc_df.head())
    # spgz_characteristics_content_loc_df = spgz_characteristics_content_loc_df[
    #     spgz_characteristics_content_loc_df['Наименование СПГЗ'].notnull() & (spgz_characteristics_content_loc_df['Наименование СПГЗ'].str.len()>0) &
    #     spgz_characteristics_content_loc_df['Наименование характеристики'].notnull() & (spgz_characteristics_content_loc_df['Наименование характеристики'].str.len()>0)
    #     ]

    unnamed_cols = [col for col in spgz_characteristics_content_loc_df.columns if col.startswith('Unnamed:')]
    spgz_characteristics_content_loc_df.drop(columns=unnamed_cols, inplace=True)

    # проверка корректности заполнения по колонкам '№п/п', 'Наименование СПГЗ', 'Единица измерения'
    # количество непустых строк по ним должно совпадать поскольку они объединены это не всегда видно визуально

    npp_nunique = spgz_characteristics_content_loc_df['№п/п'].nunique()
    mask_for_value_counts = (spgz_characteristics_content_loc_df['№п/п'].notnull() |
    spgz_characteristics_content_loc_df['Наименование СПГЗ'].notnull() |
    spgz_characteristics_content_loc_df['Единица измерения'].notnull()
    )
    # что0нибудь непустое должно быть, если все пустые колонки - не считаем
    need_value_counts = spgz_characteristics_content_loc_df[mask_for_value_counts].value_counts(['№п/п', 'Наименование СПГЗ', 'Единица измерения'], dropna=False).shape[0]
    # name_spgz_nunique = spgz_characteristics_content_loc_df['Наименование СПГЗ'].nunique()
    # ei_nunique = spgz_characteristics_content_loc_df['Единица измерения'].nunique()
    # assert ((npp_nunique!=name_spgz_nunique) and (name_spgz_nunique != ei_nunique) and  (npp_nunique!=ei_nunique),
    #         "Ошибка заполнения объединенных ячеек по колонкам '№п/п', 'Наименование СПГЗ', 'Единица измерения'"
    #         )
    if debug: print(f"npp_nunique: '{npp_nunique}', need_value_counts: '{need_value_counts}'")
    if npp_nunique!= need_value_counts:
        print("Ошибка заполнения объединенных ячеек по колонкам '№п/п', 'Наименование СПГЗ', 'Единица измерения'")
        print(spgz_characteristics_content_loc_df[mask_for_value_counts].value_counts(['№п/п', 'Наименование СПГЗ', 'Единица измерения'], dropna=False))

    spgz_characteristics_content_loc_df['CГПЗ Код Наименование'] = spgz_code_name
    # spgz_characteristics_content_loc_df.columns = ['CГПЗ Код Наименование'] + list(spgz_characteristics_content_loc_df.columns[:-1])
    spgz_characteristics_content_loc_df = spgz_characteristics_content_loc_df[['CГПЗ Код Наименование'] + list(spgz_characteristics_content_loc_df.columns[:-1])]

    for col in unique_test_cols:
        spgz_characteristics_content_loc_df['Ошибка объединения ячеек\n' + col] = None

    spgz_characteristics_content_loc_df_groupped = spgz_characteristics_content_loc_df.groupby(groupby_col)

    for group_name, group_df in spgz_characteristics_content_loc_df_groupped:
        # print(group_name)
        # # проверка корректности заполнения по колонкам '№п/п', 'Наименование СПГЗ', 'Единица измерения'
        # # количество непустых строк по ним должно совпадать поскольку они объединены это не всегда видно визуально
        # npp_len = group_df.shape[0]
        # name_spgz_len = group_df['Наименование СПГЗ'].nunique()
        # ei_nunique = group_df['Единица измерения'].nunique()
        # assert ((npp_nunique!=name_spgz_nunique) and (name_spgz_nunique != ei_nunique) and  (npp_nunique!=ei_nunique),
        #         "Ошибка заполнения объединенных ячеек по колонкам '№п/п', 'Наименование СПГЗ', 'Единица измерения'"
        #         )
        # print(f"npp_nunique: '{npp_nunique}', name_spgz_nunique: '{name_spgz_nunique}', ei_nunique: '{ei_nunique}'")
        for col in unique_test_cols:
            # print(group_df[col].unique())
            if group_df[col].nunique()>1:
                group_err_str = str(group_df[col].unique().tolist())
                print(f"'{groupby_col}': '{group_name}'", f"Ошибка группировки (объединения) - в колонке '{col}'")
                print(group_err_str)
                # display(spgz_characteristics_content_loc_df[(spgz_characteristics_content_loc_df[groupby_col]==group_name)])
                spgz_characteristics_content_loc_df.loc[
                    (spgz_characteristics_content_loc_df[groupby_col]==group_name), 'Ошибка объединения ячеек\n' + col] = group_err_str
    # display(spgz_characteristics_content_loc_df.head())
    try:
        spgz_characteristics_content_loc_df['Наименование характеристики'] = spgz_characteristics_content_loc_df[
            'Наименование характеристики'].apply(trim_right_dot_compress_spaces)
    except Exception as err:
        print("ERROR: extract_spgz_df_lst: apply(trim_right_dot_compress_spaces)")
        print(err)
    for col in significant_cols:
        try:
            spgz_characteristics_content_loc_df[col] = spgz_characteristics_content_loc_df[col].apply(
                delete_empty_rows_in_cell)
        except Exception as err:
            print(f"ERROR: extract_spgz_df_lst: apply(delete_empty_rows_in_cell): col: '{col}'")
            print(err)
    mask = (spgz_characteristics_content_loc_df['Наименование характеристики'].notnull() &
      (spgz_characteristics_content_loc_df['Наименование характеристики'].str.len()>0)
    )
    # mask = spgz_characteristics_content_loc_df[significant_cols[0]].notnull() & (spgz_characteristics_content_loc_df[significant_cols[0]].str.len()>0)
    # for col in significant_cols[1:]:
    #     mask = mask & spgz_characteristics_content_loc_df[col].notnull() & (spgz_characteristics_content_loc_df[col].str.len()>0)
    spgz_characteristics_content_loc_df = spgz_characteristics_content_loc_df[mask]

    return spgz_code_name, spgz_characteristics_content_loc_df


# def extract_spgz_df_lst(fn, sh_n_spgz, groupby_col='№п/п',
#                         unique_test_cols=['Наименование СПГЗ', 'Единица измерения', 'ОКПД 2', 'Позиция КТРУ'],
#                         significant_cols = ['Наименование характеристики', 'Единица измерения характеристики', 'Значение характеристики', 'Тип характеристики', 'Тип выбора значений характеристики заказчиком'],
#                         debug=False,
#                         ):

#     """
#     v01.02 08.05.2024
#       изменения:
#       1. после удаления зачеркнутого текста убрать пустые строки
#       2. если вся ячейки по олокнам ниже в строке удалены то не добавлять и строку
#       ['Наименование характеристики', 'Единица измерения характеристики', 'Значение характеристики', 'Тип характеристики',
#       'Тип выбора значений характеристики заказчиком']
#       3. отбор только тех строк, где 'Наименование СПГЗ' и  'Наименование характеристики' не пусты
#       4. првоерка правильности заполнения на листе СПГЗ колонок '№п/п', 'Наименование СПГЗ', 'Единица измерения'

#     """
#     spgz_header_name_loc_df = pd.read_excel(fn, sheet_name=sh_n_spgz, header=None, nrows=1 )
#     # display(kpgz_header_name_loc_df)
#     # print(spgz_header_name_loc_df.values[0,0])
#     spgz_code_name = None
#     try:
#         spgz_code_name = re.sub( 'Перечень позиций СПГЗ, относящихся к позиции КПГЗ ', '', spgz_header_name_loc_df.values[0,0]).strip()
#     except Exception as err:
#         print("ERROR: extract_spgz_df_lst:")
#         print(err)

#     spgz_characteristics_content_loc_df = pd.read_excel(fn, sheet_name=sh_n_spgz, header=1, #, skiprows=1
#     converters = {'Наименование характеристики':str, 'Единица измерения характеристики':str, 'Значение характеристики':str, 'Тип характеристики':str,
#       'Тип выбора значений характеристики заказчиком':str}
#                                                         )
#     spgz_characteristics_content_loc_df = spgz_characteristics_content_loc_df[1:]
#     # display(spgz_characteristics_content_loc_df.head())
#     # spgz_characteristics_content_loc_df = spgz_characteristics_content_loc_df[
#     #     spgz_characteristics_content_loc_df['Наименование СПГЗ'].notnull() & (spgz_characteristics_content_loc_df['Наименование СПГЗ'].str.len()>0) &
#     #     spgz_characteristics_content_loc_df['Наименование характеристики'].notnull() & (spgz_characteristics_content_loc_df['Наименование характеристики'].str.len()>0)
#     #     ]

#     unnamed_cols = [col for col in spgz_characteristics_content_loc_df.columns if col.startswith('Unnamed:')]
#     spgz_characteristics_content_loc_df.drop(columns=unnamed_cols, inplace=True)

#     # проверка корректности заполнения по колонкам '№п/п', 'Наименование СПГЗ', 'Единица измерения'
#     # количество непустых строк по ним должно совпадать поскольку они объединены это не всегда видно визуально

#     npp_nunique = spgz_characteristics_content_loc_df['№п/п'].nunique()
#     mask_for_value_counts = (spgz_characteristics_content_loc_df['№п/п'].notnull() |
#     spgz_characteristics_content_loc_df['Наименование СПГЗ'].notnull() |
#     spgz_characteristics_content_loc_df['Единица измерения'].notnull()
#     )
#     # что0нибудь непустое должно быть, если все пустые колонки - не считаем
#     need_value_counts = spgz_characteristics_content_loc_df[mask_for_value_counts].value_counts(['№п/п', 'Наименование СПГЗ', 'Единица измерения'], dropna=False).shape[0]
#     # name_spgz_nunique = spgz_characteristics_content_loc_df['Наименование СПГЗ'].nunique()
#     # ei_nunique = spgz_characteristics_content_loc_df['Единица измерения'].nunique()
#     # assert ((npp_nunique!=name_spgz_nunique) and (name_spgz_nunique != ei_nunique) and  (npp_nunique!=ei_nunique),
#     #         "Ошибка заполнения объединенных ячеек по колонкам '№п/п', 'Наименование СПГЗ', 'Единица измерения'"
#     #         )
#     if debug: print(f"npp_nunique: '{npp_nunique}', need_value_counts: '{need_value_counts}'")
#     if npp_nunique!= need_value_counts:
#         print("Ошибка заполнения объединенных ячеек по колонкам '№п/п', 'Наименование СПГЗ', 'Единица измерения'")
#         print(spgz_characteristics_content_loc_df[mask_for_value_counts].value_counts(['№п/п', 'Наименование СПГЗ', 'Единица измерения'], dropna=False))

#     spgz_characteristics_content_loc_df['CГПЗ Код Наименование'] = spgz_code_name
#     # spgz_characteristics_content_loc_df.columns = ['CГПЗ Код Наименование'] + list(spgz_characteristics_content_loc_df.columns[:-1])
#     spgz_characteristics_content_loc_df = spgz_characteristics_content_loc_df[['CГПЗ Код Наименование'] + list(spgz_characteristics_content_loc_df.columns[:-1])]

#     for col in unique_test_cols:
#         spgz_characteristics_content_loc_df['Ошибка объединения ячеек\n' + col] = None

#     spgz_characteristics_content_loc_df_groupped = spgz_characteristics_content_loc_df.groupby(groupby_col)

#     for group_name, group_df in spgz_characteristics_content_loc_df_groupped:
#         # print(group_name)
#         # # проверка корректности заполнения по колонкам '№п/п', 'Наименование СПГЗ', 'Единица измерения'
#         # # количество непустых строк по ним должно совпадать поскольку они объединены это не всегда видно визуально
#         # npp_len = group_df.shape[0]
#         # name_spgz_len = group_df['Наименование СПГЗ'].nunique()
#         # ei_nunique = group_df['Единица измерения'].nunique()
#         # assert ((npp_nunique!=name_spgz_nunique) and (name_spgz_nunique != ei_nunique) and  (npp_nunique!=ei_nunique),
#         #         "Ошибка заполнения объединенных ячеек по колонкам '№п/п', 'Наименование СПГЗ', 'Единица измерения'"
#         #         )
#         # print(f"npp_nunique: '{npp_nunique}', name_spgz_nunique: '{name_spgz_nunique}', ei_nunique: '{ei_nunique}'")
#         for col in unique_test_cols:
#             # print(group_df[col].unique())
#             if group_df[col].nunique()>1:
#                 group_err_str = str(group_df[col].unique().tolist())
#                 print(f"'{groupby_col}': '{group_name}'", f"Ошибка группировки (объединения) - в колонке '{col}'")
#                 print(group_err_str)
#                 # display(spgz_characteristics_content_loc_df[(spgz_characteristics_content_loc_df[groupby_col]==group_name)])
#                 spgz_characteristics_content_loc_df.loc[
#                     (spgz_characteristics_content_loc_df[groupby_col]==group_name), 'Ошибка объединения ячеек\n' + col] = group_err_str
#     # display(spgz_characteristics_content_loc_df.head())
#     try:
#         spgz_characteristics_content_loc_df['Наименование характеристики'] = spgz_characteristics_content_loc_df[
#             'Наименование характеристики'].apply(trim_right_dot_compress_spaces)
#     except Exception as err:
#         print("ERROR: extract_spgz_df_lst: apply(trim_right_dot_compress_spaces)")
#         print(err)
#     for col in significant_cols:
#         try:
#             spgz_characteristics_content_loc_df[col] = spgz_characteristics_content_loc_df[col].apply(
#                 delete_empty_rows_in_cell)
#         except Exception as err:
#             print(f"ERROR: extract_spgz_df_lst: apply(delete_empty_rows_in_cell): col: '{col}'")
#             print(err)
#     mask = (spgz_characteristics_content_loc_df['Наименование характеристики'].notnull() &
#       (spgz_characteristics_content_loc_df['Наименование характеристики'].str.len()>0)
#     )
#     # mask = spgz_characteristics_content_loc_df[significant_cols[0]].notnull() & (spgz_characteristics_content_loc_df[significant_cols[0]].str.len()>0)
#     # for col in significant_cols[1:]:
#     #     mask = mask & spgz_characteristics_content_loc_df[col].notnull() & (spgz_characteristics_content_loc_df[col].str.len()>0)
#     spgz_characteristics_content_loc_df = spgz_characteristics_content_loc_df[mask]

#     return spgz_code_name, spgz_characteristics_content_loc_df

def extract_spgz_df_lst_v00(fn, sh_n_spgz, groupby_col='№п/п',
                        unique_test_cols=['Наименование СПГЗ', 'Единица измерения', 'ОКПД 2', 'Позиция КТРУ']):

    spgz_header_name_loc_df = pd.read_excel(fn, sheet_name=sh_n_spgz, header=None, nrows=1 )
    # display(kpgz_header_name_loc_df)
    # print(spgz_header_name_loc_df.values[0,0])
    spgz_code_name = None
    try:
        spgz_code_name = re.sub( 'Перечень позиций СПГЗ, относящихся к позиции КПГЗ ', '', spgz_header_name_loc_df.values[0,0]).strip()
    except Exception as err:
        print("ERROR: extract_spgz_df_lst:")
        print(err)

    spgz_characteristics_content_loc_df = pd.read_excel(fn, sheet_name=sh_n_spgz, header=1) #, skiprows=1
    spgz_characteristics_content_loc_df = spgz_characteristics_content_loc_df[1:]
    # display(spgz_characteristics_content_loc_df.head())

    unnamed_cols = [col for col in spgz_characteristics_content_loc_df.columns if col.startswith('Unnamed:')]
    spgz_characteristics_content_loc_df.drop(columns=unnamed_cols, inplace=True)

    spgz_characteristics_content_loc_df['CГПЗ Код Наименование'] = spgz_code_name
    # spgz_characteristics_content_loc_df.columns = ['CГПЗ Код Наименование'] + list(spgz_characteristics_content_loc_df.columns[:-1])
    spgz_characteristics_content_loc_df = spgz_characteristics_content_loc_df[['CГПЗ Код Наименование'] + list(spgz_characteristics_content_loc_df.columns[:-1])]

    for col in unique_test_cols:
        spgz_characteristics_content_loc_df['Ошибка объединения ячеек\n' + col] = None

    spgz_characteristics_content_loc_df_groupped = spgz_characteristics_content_loc_df.groupby(groupby_col)

    for group_name, group_df in spgz_characteristics_content_loc_df_groupped:
        # print(group_name)
        for col in unique_test_cols:
            # print(group_df[col].unique())
            if group_df[col].nunique()>1:
                group_err_str = str(group_df[col].unique().tolist())
                print(f"'{groupby_col}': '{group_name}'", f"Ошибка группировки (объединения) - в колонке '{col}'")
                print(group_err_str)
                # display(spgz_characteristics_content_loc_df[(spgz_characteristics_content_loc_df[groupby_col]==group_name)])
                spgz_characteristics_content_loc_df.loc[
                    (spgz_characteristics_content_loc_df[groupby_col]==group_name), 'Ошибка объединения ячеек\n' + col] = group_err_str
    # display(spgz_characteristics_content_loc_df.head())
    try:
        spgz_characteristics_content_loc_df['Наименование характеристики'] = spgz_characteristics_content_loc_df['Наименование характеристики'].apply(trim_right_dot_compress_spaces)
    except Exception as err:
        print(err)

    return spgz_code_name, spgz_characteristics_content_loc_df

def pivot_combine_kpgz_spgz_xlsx(fn_lst):
    kpgz_header_content_df = []
    kpgz_characteristics_content_df = []
    spgz_characteristics_content_df = []

    for fn in fn_lst:
        print(fn)
        kpgz_code_name, kpgz_header_content_loc_df, kpgz_characteristics_content_loc_df = extract_kpgz_df_lst(fn, sh_n_kpgz='КПГЗ')
        # print("kpgz_code_name:", kpgz_code_name)
        # spgz_code_name, spgz_characteristics_content_loc_df = extract_spgz_df_lst(fn, sh_n_spgz='КПГЗ')
        spgz_code_name, spgz_characteristics_content_loc_df = extract_spgz_df_lst_st(fn, sh_n_spgz='КПГЗ')
        print("kpgz_code_name:", kpgz_code_name, "spgz_code_name:", spgz_code_name)
        # col = 'Наименование СПГЗ'
        # display(spgz_characteristics_content_loc_df[spgz_characteristics_content_loc_df['Ошибка объединения ячеек\n' + col].notnull()])
        kpgz_header_content_df.append(kpgz_header_content_loc_df)
        kpgz_characteristics_content_df.append(kpgz_characteristics_content_loc_df)
        spgz_characteristics_content_df.append(spgz_characteristics_content_loc_df)

        # break
    kpgz_header_content_df = pd.concat(kpgz_header_content_df)
    # display(kpgz_header_content_df.head())
    kpgz_characteristics_content_df = pd.concat(kpgz_characteristics_content_df)
    # display(kpgz_characteristics_content_df.head())
    spgz_characteristics_content_df = pd.concat(spgz_characteristics_content_df)
    # display(spgz_characteristics_content_df.head())

    return kpgz_header_content_df, kpgz_characteristics_content_df, spgz_characteristics_content_df

def get_single_value_chars_of_chars(
    chars_of_chars_dict,
    spgz_df_value_counts,
    name_char_col, name_char_of_char_col, count_col,
    default_value,
    value_def_in_spgz = 'Определено в СПГЗ',
    debug=False):

    spgz_df_value_counts_groupped = spgz_df_value_counts.drop(columns=[count_col]).groupby(name_char_col)

    for group_name, group_df in spgz_df_value_counts_groupped:
        if debug:
            print(group_name) # name_char_of_char
        if chars_of_chars_dict.get(group_name) is None:
            chars_of_chars_dict[group_name] = {}
        match (group_df.shape[0]):
            case 1:
                chars_of_chars_dict[group_name][name_char_of_char_col] = group_df.values[0,-1]
                if debug: print(chars_of_chars_dict[group_name])
            case 0:
                # chars_of_chars_dict[group_name][name_char_of_char_col] = default_value
                """
                (1) Если нет столбца «Условная операция», выполнять следующее:
                если «Значение характеристики» содержит символы >= и/ или <=,
                то в поле КПГЗ «Условная операция» указать «Диапазон», иначе указать «Перечисление»
                """
                if group_name=='Условная операция': # name_char_of_char
                    if (
                        '>=' in  chars_of_chars_dict[group_name]['Значение характеристики'] or
                        '<=' in  chars_of_chars_dict[group_name]['Значение характеристики']
                    ):
                        chars_of_chars_dict[group_name][name_char_of_char_col] = 'Диапазон'
                    else:
                        chars_of_chars_dict[group_name][name_char_of_char_col] = 'Перечисление'
            case _:
                chars_of_chars_dict[group_name][name_char_of_char_col] = value_def_in_spgz
    # if spgz_df_value_counts[spgz_df_value_counts.duplicated(name_char_col)].shape[0] > 0:
    #     print("duplicated:")
    #     pprint(spgz_df_value_counts[spgz_df_value_counts.duplicated(name_char_col)][name_char_col].drop_duplicates().values)
    # else:
    #     print("no duplicates!!!")

    return chars_of_chars_dict

def get_joined_chars_of_chars(
    chars_of_chars_dict,
    spgz_df_value_counts, sep,
    name_char_col, name_char_of_char_col, count_col,
                              debug= False):


    spgz_df_value_counts_groupped = spgz_df_value_counts.drop(columns=[count_col]).groupby(name_char_col)

    for group_name, group_df in spgz_df_value_counts_groupped:
        if debug:
            print(group_name) # name_char_of_char
        if chars_of_chars_dict.get(group_name) is None:
            chars_of_chars_dict[group_name] = {}
        chars_of_chars_dict[group_name][name_char_of_char_col] = None
        group_values_lst = group_df.values[:,-1]
        if debug:
            print("до переразбиения:")
            print(group_values_lst)
        # перазбиваем полученные значения
        # Типа ['>=1.5<3\n<=3', '>=1.5']
        group_values_lst = [v.split(sep) for v in group_values_lst]
        group_values_lst = [v_element for v_lst in group_values_lst for v_element in v_lst]
        group_values_lst = sorted(list(set(group_values_lst)))
        # if debug:
        #     print("после переразбиения:")
        #     print(group_values_lst)
        # Обработка ситуации исключения дублей
        #  ['Взрослые', 'Дети', 'Новорожденные', 'взрослые', 'дети', 'новорожденные']
        group_values_set = sorted(set([v.lower() for v in group_values_lst]), key=len, reverse=True)
        if len(set(group_values_set)) < len(group_values_lst):
            group_values_lst = sorted(list_drop_duplicates(group_values_lst, sep))
            # if debug:
            #     print("после удаления дублей:")
            #     print(group_values_lst)
        group_values_str = sep.join(group_values_lst)
        chars_of_chars_dict[group_name][name_char_of_char_col] = group_values_str
        if debug:
            print("после переразбиения:")
            print(group_values_lst)
    return chars_of_chars_dict

def list_drop_duplicates(group_values_lst, sep, debug=False):
    # ['Взрослые', 'Дети', 'Новорожденные', 'взрослые', 'дети', 'новорожденные']
    # sep для того чтобы не было пересейчений с возможными другими разделителями
    group_values_set = sorted(set([v.lower() for v in group_values_lst]), key=len, reverse=True)
    # сортировка, чтобы удалялиьс сначала наиболее длинные слова, чтобы корткие слова, к-ые могут входить в более длинные слова своим удалением не разрушили общую строку
    if debug: print(group_values_set)
    if len(group_values_set)< len(group_values_lst):
        group_values_str = sep.join(group_values_lst)
        # print(group_values_str)
        # преимущественно удалим lower
        group_values_set_tmp = []
        for vv in group_values_set:
            group_values_str = re.sub(re.escape(vv), '', group_values_str, count=1)
            group_values_set_tmp.append(vv)
        group_values_lst_tmp = group_values_str.split(sep)
        group_values_str = sep.join(group_values_lst_tmp)
        if len(group_values_lst_tmp) > len(group_values_set):
            # group_values_str = sep.join(group_values_lst_tmp)
            for vv in group_values_set:
                if vv not in group_values_set_tmp:
                    group_values_str = re.sub(re.escape(vv), '', group_values_str, flags=re.I, count=1)
        group_values_lst_new = group_values_str.split(sep)
        group_values_lst_new = [v for v in group_values_lst_new if len(v) > 0]
        return group_values_lst_new
    else:
        return group_values_lst

def create_kpgz_data_01(
    spgz_characteristics_content_loc_df, debug = False,
):
    spgz_df = spgz_characteristics_content_loc_df.copy()
    cols_for_kpgz_head = {
        'Характеристики (кол-во)': 'Наименование характеристики',
        'СПГЗ (кол-во)': 'Наименование СПГЗ',
        'СПГЗ/ИНП (кол-во)': ['Наименование СПГЗ', 'ИНП'],
        'ОКПД-2': 'ОКПД 2',
        # 'Код характеристики КТРУ': 'Позиция КТРУ',
        'Позиция КТРУ': 'Позиция КТРУ',
    }
    kpgz_head_indicators_type = {
        'Характеристики (кол-во)': 'Количество',
        'СПГЗ/ИНП (кол-во)': 'Количество',
        'СПГЗ (кол-во)': 'Количество',
        'ОКПД-2': 'Перечень',
        'Позиция КТРУ': 'Перечень',}

    kpgz_head = {}

    name_char_col='Наименование характеристики'
    count_col = 'count'
    char_of_char_cols_lst = [
        'Единица измерения характеристики',
        'Значение характеристики',
        'Тип характеристики',
        'Условная операция',
        'Тип выбора значений характеристики заказчиком',
        'Позиция КТРУ',
    ]
    sep_lst = {
        'Единица измерения характеристики': ';\n',
        'Значение характеристики': '\n',
    }
    default_value_lst = {
        'Условная операция': 'Перечисление',
    }

    value_def_in_spgz = 'Определено в СПГЗ'

    missing_columns = []

    # 'Определено в СПГЗ'
    for sub_kpgz_head, cols in cols_for_kpgz_head.items():
        if (
            ((type(cols)==list) and set(cols).issubset(set(spgz_df.columns))) or
            ((type(cols)==str) and cols in spgz_df.columns)
        ):

            if kpgz_head_indicators_type[sub_kpgz_head]=='Количество':
                if type(cols)==str:
                    kpgz_head[sub_kpgz_head] = spgz_df[cols].nunique()
                    # print(f"'{cols}'", spgz_df[cols].nunique())
                elif type(cols)==list:
                    kpgz_head[sub_kpgz_head] = spgz_df.value_counts(cols).shape[0]
                    # print(f"'{cols}'", spgz_df.value_counts(cols).shape[0])
            elif kpgz_head_indicators_type[sub_kpgz_head]=='Перечень':
                    kpgz_head[sub_kpgz_head] = sorted(spgz_df[cols].unique())
                    # только для одиночных колонок
                    # print(spgz_df[cols].unique()[:5])
        else:
            print(f"Колонка/колонки '{cols}' отсутствуют в данных")
    if debug:
        print("kpgz_head:")
        pprint(kpgz_head)
        print(80*'*')

    chars_of_chars_dict = {}
    for col in char_of_char_cols_lst:
        name_char_of_char_col = col
        if col in spgz_df.columns:

            value_counts_cols_pair = [name_char_col, col]
            if debug: print(f"{value_counts_cols_pair}")
            spgz_df_value_counts = spgz_df.value_counts(value_counts_cols_pair).reset_index().sort_values(value_counts_cols_pair)
            # display(spgz_df_value_counts)
            if sep_lst.get(name_char_of_char_col) is not None:
                sep = sep_lst.get(name_char_of_char_col)
                # sep_test_01 = sep
                # spgz_df_value_counts_test_01 = spgz_df_value_counts.copy()
                # name_char_of_char_col_test_01 = col
                # display(spgz_df_value_counts.head(5))
                # break

                chars_of_chars_dict = get_joined_chars_of_chars(
                    chars_of_chars_dict,
                    spgz_df_value_counts, sep,
                    name_char_col, name_char_of_char_col, count_col,
                                              debug=False)
            else:
                chars_of_chars_dict = get_single_value_chars_of_chars(
                    chars_of_chars_dict,
                    spgz_df_value_counts,
                    name_char_col, name_char_of_char_col, count_col,
                    default_value=default_value_lst.get(name_char_of_char_col),
                    value_def_in_spgz = value_def_in_spgz, #'Определено в СПГЗ',
                    debug=False)
        else:
            logger.error(f"Колонка '{name_char_of_char_col}' отсутствует в данных")
            missing_columns.append(name_char_of_char_col)
            for key, value_dict in chars_of_chars_dict.items():
                chars_of_chars_dict[key][name_char_of_char_col] = default_value_lst.get(name_char_of_char_col)

    # name_char_col='Наименование характеристики'
    chars_of_chars_dict_lst = []
    for key, value_dict in chars_of_chars_dict.items():
        loc_dict = {name_char_col: key}
        loc_dict.update(value_dict)
        chars_of_chars_dict_lst.append(loc_dict)

    # chars_of_chars_dict_lst[:5]
    chars_of_chars_df = pd.DataFrame(chars_of_chars_dict_lst)
    chars_of_chars_df.rename(columns ={'Позиция КТРУ': 'Код характеристики КТРУ'}, inplace=True)
    # chars_of_chars_df.head()

    return kpgz_head, chars_of_chars_df

def create_kpgz_data(
    spgz_characteristics_content_loc_df, debug = False,
):
    """
    v01.02 21.10.2024
        bag_fix
            'Код КТРУ характеристики',
    """
    spgz_df = spgz_characteristics_content_loc_df.copy()
    cols_for_kpgz_head = {
        'Характеристики (кол-во)': 'Наименование характеристики',
        'СПГЗ (кол-во)': 'Наименование СПГЗ',
        'СПГЗ/ИНП (кол-во)': ['Наименование СПГЗ', 'ИНП'],
        'ОКПД-2': 'ОКПД 2',
        # 'Код характеристики КТРУ': 'Позиция КТРУ',
        'Позиция КТРУ': 'Позиция КТРУ',
    }
    kpgz_head_indicators_type = {
        'Характеристики (кол-во)': 'Количество',
        'СПГЗ/ИНП (кол-во)': 'Количество',
        'СПГЗ (кол-во)': 'Количество',
        'ОКПД-2': 'Перечень',
        'Позиция КТРУ': 'Перечень',}

    kpgz_head = {}

    name_char_col='Наименование характеристики'
    count_col = 'count'
    char_of_char_cols_lst = [
        'Единица измерения характеристики',
        'Значение характеристики',
        'Тип характеристики',
        'Условная операция',
        'Тип выбора значений характеристики заказчиком',
        # 'Позиция КТРУ',
        # 'Код характеристики КТРУ',
        'Код КТРУ характеристики',
    ]
    sep_lst = {
        'Единица измерения характеристики': ';\n',
        'Значение характеристики': '\n',
    }
    default_value_lst = {
        'Условная операция': 'Перечисление',
    }

    value_def_in_spgz = 'Определено в СПГЗ'

    missing_columns = []

    # 'Определено в СПГЗ'
    for sub_kpgz_head, cols in cols_for_kpgz_head.items():
        if (
            ((type(cols)==list) and set(cols).issubset(set(spgz_df.columns))) or
            ((type(cols)==str) and cols in spgz_df.columns)
        ):

            if kpgz_head_indicators_type[sub_kpgz_head]=='Количество':
                if type(cols)==str:
                    kpgz_head[sub_kpgz_head] = spgz_df[cols].nunique()
                    # print(f"'{cols}'", spgz_df[cols].nunique())
                elif type(cols)==list:
                    kpgz_head[sub_kpgz_head] = spgz_df.value_counts(cols).shape[0]
                    # print(f"'{cols}'", spgz_df.value_counts(cols).shape[0])
            elif kpgz_head_indicators_type[sub_kpgz_head]=='Перечень':
                    kpgz_head[sub_kpgz_head] = sorted(spgz_df[cols].unique())
                    # только для одиночных колонок
                    # print(spgz_df[cols].unique()[:5])
        else:
            print(f"Колонка/колонки '{cols}' отсутствуют в данных (Заголовок)")
            print(spgz_df.columns)
    if debug:
        print("kpgz_head:")
        pprint(kpgz_head)
        print(80*'*')

    chars_of_chars_dict = {}
    for col in char_of_char_cols_lst:
        name_char_of_char_col = col
        # st.write(f"'{col}'")
        if col in spgz_df.columns:

            value_counts_cols_pair = [name_char_col, col]
            if debug: print(f"{value_counts_cols_pair}")
            spgz_df_value_counts = spgz_df.value_counts(value_counts_cols_pair).reset_index().sort_values(value_counts_cols_pair)
            # display(spgz_df_value_counts)
            if sep_lst.get(name_char_of_char_col) is not None:
                sep = sep_lst.get(name_char_of_char_col)
                # sep_test_01 = sep
                # spgz_df_value_counts_test_01 = spgz_df_value_counts.copy()
                # name_char_of_char_col_test_01 = col
                # display(spgz_df_value_counts.head(5))
                # break

                chars_of_chars_dict = get_joined_chars_of_chars(
                    chars_of_chars_dict,
                    spgz_df_value_counts, sep,
                    name_char_col, name_char_of_char_col, count_col,
                                              debug=False)
            else:
                chars_of_chars_dict = get_single_value_chars_of_chars(
                    chars_of_chars_dict,
                    spgz_df_value_counts,
                    name_char_col, name_char_of_char_col, count_col,
                    default_value=default_value_lst.get(name_char_of_char_col),
                    value_def_in_spgz = value_def_in_spgz, #'Определено в СПГЗ',
                    debug=False)
        else:
            # logger.error(f"Колонка '{name_char_of_char_col}' отсутствует в данных (Таблица)")
            st.write(f"Колонка '{name_char_of_char_col}' отсутствует в данных (Таблица)")
            # print(spgz_df.columns)
            missing_columns.append(name_char_of_char_col)
            for key, value_dict in chars_of_chars_dict.items():
                chars_of_chars_dict[key][name_char_of_char_col] = default_value_lst.get(name_char_of_char_col)

    # name_char_col='Наименование характеристики'
    chars_of_chars_dict_lst = []
    for key, value_dict in chars_of_chars_dict.items():
        loc_dict = {name_char_col: key}
        loc_dict.update(value_dict)
        chars_of_chars_dict_lst.append(loc_dict)

    # chars_of_chars_dict_lst[:5]
    chars_of_chars_df = pd.DataFrame(chars_of_chars_dict_lst)
    try:
        chars_of_chars_df.rename(columns ={'Код КТРУ характеристики': 'Код характеристики КТРУ'}, inplace=True)
    except:
        pass

    # chars_of_chars_df.head()
    # st.write(f"create_kpgz_data: return kpgz_head, chars_of_chars_df")
    return kpgz_head, chars_of_chars_df

def get_total_okpd2_code_name(
    kpgz_head,
    okpd2_df,
    debug=False,

):
    """
    v01.02 21.10.2024
      изменения:
      если ОКПД-2 все одинаковые, то подгружаем полностью, если ОКПД-2 несколько, то подгружаем только первые 5 символов (до первого отличающегося символа)
    """
    if debug: print(kpgz_head['ОКПД-2'])
    okpd2_lst = kpgz_head['ОКПД-2']
    if okpd2_lst is not None and (len(okpd2_lst) > 0):
        okpd2_lst_upd = [code_name.replace('\n','').strip() for code_name in okpd2_lst]
        if '-' in okpd2_lst_upd:
            return '-'
        # okpd2_codes_lst = [re.sub(r"(?:[^\d\.]+)", '', s) for s in okpd2_lst_upd]
        # okpd2_codes_lst = [re.sub(r"(?:[^\d\.]+)(?=\s)", '', s) for s in okpd2_lst_upd]
        okpd2_codes_lst = [re.search(r"^[\d\.]+", s).group(0) for s in okpd2_lst_upd if re.search(r"^[\d\.]+", s) is not None]
        if debug: print(okpd2_codes_lst)
        set_okpd2_codes_lst = set(okpd2_codes_lst)
        if len(set_okpd2_codes_lst)==1:
            okpd2_code_prefix = okpd2_codes_lst[0]
            if debug: print(okpd2_code_prefix)
        elif len(set_okpd2_codes_lst) > 1:
            okpd2_code_prefix = os.path.commonprefix(okpd2_codes_lst)
            if debug: print(okpd2_code_prefix)
        else: # ==0
            return '-'

        okpd2_code_prefix = clean_str(okpd2_code_prefix)
        if debug: print(f"okpd2_code_prefix: '{okpd2_code_prefix}'")
        # ['КОД РАЗДЕЛА', 'НАИМЕНОВАНИЕ РАЗДЕЛА', 'КОД ПОЗИЦИИ', 'НАИМЕНОВАНИЕ ПОЗИЦИИ']
        # ['Код', 'Наименование', 'Комментарий', 'Тип кода']
        okpd2_code_name_lst = okpd2_df[okpd2_df['Код']==okpd2_code_prefix]['Наименование'].values
        if debug: print("okpd2_code_name_lst:", okpd2_code_name_lst)
        if len(okpd2_code_name_lst) > 0:
            return okpd2_code_prefix + ' ' + okpd2_code_name_lst[0]
        else:
            return okpd2_code_prefix

def get_total_ktru_code_name(
    kpgz_head,
    ktru_obj_name = 'Позиция КТРУ',
    # ktru_df,
    sep='|\n',
    debug=False,

):
    """
    v01.02 21.10.2024
         изменения:
          если ОКПД-2 все одинаковые, то подгружаем полностью, если ОКПД-2 несколько, то подгружаем только первые 5 символов (до первого отличающегося символа)

    """
    if debug: print(kpgz_head[ktru_obj_name])
    ktru_lst = kpgz_head[ktru_obj_name]
    ktru_is_lst = False

    if ktru_lst is not None and (len(ktru_lst) > 0):
        ktru_lst_upd = [ktru_name.replace('\n','').strip() for ktru_name in ktru_lst]
        if '-' in ktru_lst_upd:
            return '-', ktru_is_lst
        # else:
        #     return sep.join(ktru_lst_upd), True
        # ktru_codes_lst = [re.search(r"^[\d\.]+", s).group(0) for s in ktru_lst_upd if re.search(r"^[\d\.]+", s) is not None]
        ktru_codes_lst = [re.search(r"^[\d\.\-]+", s).group(0) for s in ktru_lst_upd if re.search(r"^[\d\.\-]+", s) is not None]
        if debug: print(ktru_codes_lst)
        set_ktru_codes_lst = set(ktru_codes_lst)
        if debug: print(set_ktru_codes_lst)

        if len(set_ktru_codes_lst)==1:
            # ktru_code_prefix = ktru_codes_lst[0]
            ktru_code_prefix = ktru_lst_upd[0]
            if debug: 
                print("len(set_ktru_codes_lst)==1")
                print(ktru_code_prefix)
                
            return ktru_code_prefix, ktru_is_lst
        elif len(set_ktru_codes_lst) > 1:
            # ktru_code_prefix = os.path.commonprefix(ktru_codes_lst)
            ktru_code_prefix = os.path.commonprefix(ktru_lst)
            ktru_is_lst = True
            if debug: 
                print("len(set_ktru_codes_lst)>1")
                print(ktru_code_prefix)
            return ktru_code_prefix, ktru_is_lst
        else: # ==0
            return '-', ktru_is_lst

        # ktru_code_prefix = clean_str(ktru_code_prefix)
        # if debug: print(f"ktru_code_prefix: '{ktru_code_prefix}'")
    else: # ktru_lst is None or (len(ktru_lst) == 0):
        return '-', ktru_is_lst

# def get_total_ktru_code_name_v01(
#     kpgz_head,
#     ktru_obj_name = 'Позиция КТРУ',
#     # ktru_df,
#     sep='|\n',
#     debug=False,

# ):
#     if debug: print(kpgz_head[ktru_obj_name])
#     ktru_lst = kpgz_head[ktru_obj_name]
#     ktru_is_lst = False

#     if ktru_lst is not None and (len(ktru_lst) > 0):
#         ktru_lst_upd = [ktru_name.replace('\n','').strip() for ktru_name in ktru_lst]
#         if '-' in ktru_lst_upd:
#             return '-', ktru_is_lst
#         # else:
#         #     return sep.join(ktru_lst_upd), True
#         ktru_codes_lst = [re.search(r"^[\d\.]+", s).group(0) for s in ktru_lst_upd if re.search(r"^[\d\.]+", s) is not None]
#         if debug: print(ktru_codes_lst)
#         set_ktru_codes_lst = set(ktru_codes_lst)
#         print(set_ktru_codes_lst)
#         if len(set_ktru_codes_lst)==1:
#             ktru_code_prefix = ktru_codes_lst[0]
#             if debug: print(ktru_code_prefix)
#         elif len(set_ktru_codes_lst) > 1:
#             ktru_code_prefix = os.path.commonprefix(ktru_codes_lst)
#             ktru_is_lst = True
#             if debug: print(ktru_code_prefix)
#         else: # ==0
#             return '-', ktru_is_lst

#         ktru_code_prefix = clean_str(ktru_code_prefix)
#         if debug: print(f"ktru_code_prefix: '{ktru_code_prefix}'")

#         return ktru_lst, ktru_is_lst

# ktru_lst, ktru_is_lst = get_total_ktru_code_name(
#     kpgz_head,
#     # okpd2_df,
#     debug=True)


from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from copy import copy

def write_head_kpgz_sheet_st_v01(
    fn_source,
    fn_save,
    spgz_code_name,
    kpgz_head,
    chars_of_chars_df,
    okpd2_df,
    debug=False
 ):
    column_widths = [40,20,60,20,20,25,25,]
    ft_bold = Font(bold = True)
    ft_norm = Font(bold = False)
    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
    # border = Border(left=Side(border_style=None,
    #                       color='FF000000'),
    #             right=Side(border_style=None,
    #                        color='FF000000'),
    #             top=Side(border_style=None,
    #                      color='FF000000'),
    #             bottom=Side(border_style=None,
    #                         color='FF000000'),
    #             diagonal=Side(border_style=None,
    #                           color='FF000000'),
    #             diagonal_direction=0,
    #             outline=Side(border_style=None,
    #                          color='FF000000')

    wb = Workbook()
    ws = wb.active
    ws.title = 'КПГЗ'

    ws['A1'] = 'Позиция КПГЗ ' + spgz_code_name
    ws['A1'].font = ft_bold
    ws['A1'].border = thin_border
    ws['A1'].alignment = Alignment(wrap_text=True,vertical='top', horizontal='center')

    ws['A2'] = 'Статус'
    ws['A2'].font = ft_bold
    ws['B2'] = 'Утверждена'
    ws['A2'].border = thin_border
    ws['B2'].border = thin_border

    ws['A3'] = 'ОКПД-2'
    ws['A3'].font = ft_bold
    # ws['B3'] = kpgz_head['ОКПД-2'][0]  #okpd2
    ws['B3'] = get_total_okpd2_code_name(
                kpgz_head,
                okpd2_df,
                debug=False)
    ws['A3'].border = thin_border
    ws['B3'].border = thin_border

    ws['A4'] = 'Позиция КТРУ'
    ws['A4'].font = ft_bold
    ws['A4'].border = thin_border
    ws['B4'].border = thin_border

    ktru_lst, ktru_is_lst = get_total_ktru_code_name(kpgz_head)
    ws['B4'] = ktru_lst
    if ktru_is_lst and (type(ktru_lst)!=str):
        ws['B4'].fill =  PatternFill('solid', fgColor='00C0C0C0')

    ws['A5'] = 'Уровень детализации адреса'
    ws['A5'].font = ft_bold
    ws['B5'] = '-'
    ws['A5'].border = thin_border
    ws['B5'].border = thin_border

    ws['A6'] = 'Загружено из ЕМИАС'
    ws['A6'].font = ft_bold
    ws['B6'] = 'Нет'
    ws['A6'].border = thin_border
    ws['B6'].border = thin_border

    ws['A7'] = 'Характеристики (кол-во)'
    ws['A7'].font = ft_bold
    ws['B7'] = str(kpgz_head['Характеристики (кол-во)'])
    ws['A7'].border = thin_border
    ws['B7'].border = thin_border

    ws['A8'] = 'СПГЗ (кол-во)'
    ws['A8'].font = ft_bold
    ws['B8'] = f"{kpgz_head['СПГЗ (кол-во)']} (СПГЗ/ИНП (кол-во) {kpgz_head['СПГЗ/ИНП (кол-во)']})"
    ws['A8'].border = thin_border
    ws['B8'].border = thin_border


    ws['A10'] = 'Справочник характеристик и их значений позиции КПГЗ ' + spgz_code_name
    ws['A10'].font = ft_bold
    ws['A10'].border = thin_border
    ws['B10'].border = thin_border
    ws['A10'].alignment = Alignment(wrap_text=True,vertical='top', horizontal='center')

    # ws.append([None])
    ws.append(list(chars_of_chars_df.columns))
    i_row = 11
    for i in range(1, len(column_widths)+1):  # ,1 to start at 1
        ws[get_column_letter(i) + f"{i_row}"].font = ft_bold
        ws[get_column_letter(i) + f"{i_row}"].alignment = Alignment(wrap_text=True,vertical='top', horizontal='center')
        ws[get_column_letter(i) + f"{i_row}"].border = thin_border

    ws.append(list(range(1,len(chars_of_chars_df.columns) + 2)))
    i_row = 12
    for i in range(1, len(column_widths)+1):  # ,1 to start at 1
        ws[get_column_letter(i) + f"{i_row}"].font = ft_bold
        ws[get_column_letter(i) + f"{i_row}"].alignment = Alignment(wrap_text=True,vertical='top', horizontal='center')
        ws[get_column_letter(i) + f"{i_row}"].border = thin_border

    i_row = 13
    # for row in chars_of_chars_df.iterrows():
    for row in chars_of_chars_df.itertuples():
        ws.append(list(row)[1:] )
        for i in range(1, len(column_widths)+1):  # ,1 to start at 1
            ws[get_column_letter(i) + f"{i_row}"].alignment = Alignment(wrap_text=True,vertical='top')
            # if i < (len(column_widths)+2):
            ws[get_column_letter(i) + f"{i_row}"].border = thin_border
        i_row += 1

    for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
        ws.column_dimensions[get_column_letter(i)].width = column_width
        # ws.column.alignment = Alignment(wrap_text=True,vertical='top')
        # ws.column_dimensions[get_column_letter(i)].alignment = Alignment(wrap_text=True,vertical='top')

    ws.merge_cells('A10:G10')
    ws.merge_cells('A1:G1')
    ws.merge_cells('B2:G2')
    ws.merge_cells('B3:G3')
    ws.merge_cells('B4:G4')
    ws.merge_cells('B5:G5')
    ws.merge_cells('B6:G6')
    ws.merge_cells('B7:G7')
    ws.merge_cells('B8:G8')

    ws_target = wb.create_sheet('СПГЗ')
    # ws_target = wb_source.copy_worksheet(ws_source) # Не работает
    # ws_target = ws_source.rows # Не работает

    # wb.save(os.path.join(data_processed_dir, fn_save))

    wb_source = load_workbook(fn_source) #, read_only=False)
    ws_source = wb_source['СПГЗ']

    for i in range(1, ws_source.max_column + 1):
        ws_target.column_dimensions[get_column_letter(i)].width = ws_source.column_dimensions[get_column_letter(i)].width
    for row in ws_source.iter_rows():
        # ws_target.append(row) # не работает
        for cell in row:
            ws_target[cell.coordinate ] = cell.value
            ws_target[cell.coordinate ].alignment = copy(cell.alignment)
            ws_target[cell.coordinate ].font = copy(cell.font)
            ws_target[cell.coordinate ].border = copy(cell.border)

    for merged_cells_range in sorted(ws_source.merged_cells.ranges):
        start_col, start_row, end_col, end_row  = merged_cells_range.bounds
        ws_target.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

    wb.save(fn_save)

    # logger.info(f"Файл '{fn_save}' - сохранен в папке '{data_processed_dir}'")
    st.write(f"Обработанный файл '{fn_save}' - сохранен!")

    return

def write_head_kpgz_sheet_st(
    fn_source,
    fn_save,
    spgz_code_name,
    kpgz_head,
    chars_of_chars_df,
    okpd2_df,
    debug=False
 )
    column_widths = [40,20,60,20,20,25,25,]
    ft_bold = Font(bold = True, name='Times New Roman')
    ft_norm = Font(bold = False, name='Times New Roman')
    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
    # border = Border(left=Side(border_style=None,
    #                       color='FF000000'),
    #             right=Side(border_style=None,
    #                        color='FF000000'),
    #             top=Side(border_style=None,
    #                      color='FF000000'),
    #             bottom=Side(border_style=None,
    #                         color='FF000000'),
    #             diagonal=Side(border_style=None,
    #                           color='FF000000'),
    #             diagonal_direction=0,
    #             outline=Side(border_style=None,
    #                          color='FF000000')

    wb = Workbook()
    ws = wb.active
    ws.title = 'КПГЗ'
    st.write(f"write_head_kpgz_sheet_st: Заголовок начало записи")
    ws['A1'] = 'Позиция КПГЗ ' + spgz_code_name
    ws['A1'].font = ft_bold
    ws['A1'].border = thin_border
    ws['A1'].alignment = Alignment(wrap_text=True,vertical='top', horizontal='center')

    ws['A2'] = 'Статус'
    ws['A2'].font = ft_bold
    ws['B2'] = 'Утверждена'
    ws['A2'].border = thin_border
    ws['B2'].border = thin_border
    ws['B2'].font = ft_norm

    ws['A3'] = 'ОКПД-2'
    ws['A3'].font = ft_bold
    # ws['B3'] = kpgz_head['ОКПД-2'][0]  #okpd2
    ws['B3'] = get_total_okpd2_code_name(
                kpgz_head,
                okpd2_df,
                debug=False)
    ws['A3'].border = thin_border
    ws['B3'].border = thin_border
    ws['B3'].font = ft_norm

    st.write(f"write_head_kpgz_sheet_st: Заголовок: Записан ОКПД2")

    try:
        ws['A4'] = 'Позиция КТРУ'
        ws['A4'].font = ft_bold
        ws['A4'].border = thin_border
        ws['B4'].border = thin_border
        ws['B4'].font = ft_norm

        ktru_lst, ktru_is_lst = get_total_ktru_code_name(kpgz_head)
        ws['B4'] = ktru_lst
        if ktru_is_lst and (type(ktru_lst)!=str):
            ws['B4'].fill =  PatternFill('solid', fgColor='00C0C0C0')
        st.write(f"write_head_kpgz_sheet_st: Заголовок: Записан КТРУ")
    except Exception as err:
        st.write(str(err))

    ws['A5'] = 'Уровень детализации адреса'
    ws['A5'].font = ft_bold
    ws['B5'] = '-'
    ws['A5'].border = thin_border
    ws['B5'].border = thin_border
    ws['B5'].font = ft_norm

    ws['A6'] = 'Загружено из ЕМИАС'
    ws['A6'].font = ft_bold
    ws['B6'] = 'Нет'
    ws['A6'].border = thin_border
    ws['B6'].border = thin_border
    ws['B6'].font = ft_norm

    ws['A7'] = 'Характеристики (кол-во)'
    ws['A7'].font = ft_bold
    ws['B7'] = str(kpgz_head['Характеристики (кол-во)'])
    ws['A7'].border = thin_border
    ws['B7'].border = thin_border
    ws['B7'].font = ft_norm

    ws['A8'] = 'СПГЗ (кол-во)'
    ws['A8'].font = ft_bold
    # ws['B8'] = f"{kpgz_head['СПГЗ (кол-во)']} (СПГЗ/ИНП (кол-во) {kpgz_head['СПГЗ/ИНП (кол-во)']})"
    ws['B8'] = f"{kpgz_head['СПГЗ/ИНП (кол-во)']}"
    ws['A8'].border = thin_border
    ws['B8'].border = thin_border
    ws['B8'].font = ft_norm


    ws['A10'] = 'Справочник характеристик и их значений позиции КПГЗ ' + spgz_code_name
    ws['A10'].font = ft_bold
    ws['A10'].border = thin_border
    ws['B10'].border = thin_border
    ws['A10'].alignment = Alignment(wrap_text=True,vertical='top', horizontal='center')

    st.write(f"write_head_kpgz_sheet_st: Заголовок записан")
    # ws.append([None])
    ws.append(list(chars_of_chars_df.columns))
    i_row = 11
    for i in range(1, len(column_widths)+1):  # ,1 to start at 1
        ws[get_column_letter(i) + f"{i_row}"].font = ft_bold
        ws[get_column_letter(i) + f"{i_row}"].alignment = Alignment(wrap_text=True,vertical='top', horizontal='center')
        ws[get_column_letter(i) + f"{i_row}"].border = thin_border

    ws.append(list(range(1,len(chars_of_chars_df.columns) + 2)))
    i_row = 12
    for i in range(1, len(column_widths)+1):  # ,1 to start at 1
        ws[get_column_letter(i) + f"{i_row}"].font = ft_bold
        ws[get_column_letter(i) + f"{i_row}"].alignment = Alignment(wrap_text=True,vertical='top', horizontal='center')
        ws[get_column_letter(i) + f"{i_row}"].border = thin_border

    i_row = 13
    # for row in chars_of_chars_df.iterrows():
    for row in chars_of_chars_df.itertuples():
        ws.append(list(row)[1:] )
        for i in range(1, len(column_widths)+1):  # ,1 to start at 1
            ws[get_column_letter(i) + f"{i_row}"].alignment = Alignment(wrap_text=True,vertical='top')
            # if i < (len(column_widths)+2):
            ws[get_column_letter(i) + f"{i_row}"].border = thin_border
            ws[get_column_letter(i) + f"{i_row}"].font = ft_norm
        i_row += 1

    for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
    # for i, column_width in enumerate(column_widths-1,1):  # ,1 to start at 1
        ws.column_dimensions[get_column_letter(i)].width = column_width
        ws.column_dimensions[get_column_letter(i)].font = Font(name='Times New Roman')
        # ws.column.alignment = Alignment(wrap_text=True,vertical='top')
        # ws.column_dimensions[get_column_letter(i)].alignment = Alignment(wrap_text=True,vertical='top')

    ws.merge_cells('A10:G10')
    ws.merge_cells('A1:G1')
    ws.merge_cells('B2:G2')
    ws.merge_cells('B3:G3')
    ws.merge_cells('B4:G4')
    ws.merge_cells('B5:G5')
    ws.merge_cells('B6:G6')
    ws.merge_cells('B7:G7')
    ws.merge_cells('B8:G8')

    ws_target = wb.create_sheet('СПГЗ')
    # ws_target = wb_source.copy_worksheet(ws_source) # Не работает
    # ws_target = ws_source.rows # Не работает

    # wb.save(os.path.join(data_processed_dir, fn_save))

    # wb_source = load_workbook(filename=os.path.join(data_source_dir, fn_source)) #, read_only=False)
    wb_source = load_workbook(filename=fn_source) #, read_only=False)
    ws_source = wb_source['СПГЗ']

    for i in range(1, ws_source.max_column + 1):
        ws_target.column_dimensions[get_column_letter(i)].width = ws_source.column_dimensions[get_column_letter(i)].width
    for row in ws_source.iter_rows():
        # ws_target.append(row) # не работает
        for cell in row:
            ws_target[cell.coordinate ] = cell.value
            ws_target[cell.coordinate ].alignment = copy(cell.alignment)
            ws_target[cell.coordinate ].font = copy(cell.font)
            ws_target[cell.coordinate ].border = copy(cell.border)

    for merged_cells_range in sorted(ws_source.merged_cells.ranges):
        start_col, start_row, end_col, end_row  = merged_cells_range.bounds
        ws_target.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

    # wb.save(os.path.join(data_processed_dir, fn_save))
    wb.save(fn_save)

    # logger.info(f"Файл '{fn_save}' - сохранен в папке '{data_processed_dir}'")
    st.write(f"Файл '{fn_save}' - сохранен в папке '{data_processed_dir}'")

    return


def main_03(
    sh_n_source = 'СПГЗ',
    debug=False,
):

    okpd2_df = read_okpd_dict_fr_link()
    st.dataframe(okpd2_df.head(2)) #, use_container_width=True)

    uploaded_files = st.file_uploader(
        "Загрузите xlsx-файлы для обработки", accept_multiple_files=True
    )
    if uploaded_files:
        fn_lst = [fn.name for fn in  uploaded_files if fn.name.endswith('.xlsx')]
        fn_save_lst = []
        if len (fn_lst) == 0:
            st.write(f"В загруженных файлах не найдены .xlsx файлы")
            # st.write(f"Работа программы завершена")
            # st.write(f"Обновите страницу")
            # sys.exit(2)
        else:
            fn_save_lst = []
            for uploaded_file in uploaded_files:
                if uploaded_file.name.endswith('.xlsx'):
                    fn_proc_save = split_merged_cells_st(uploaded_file, sh_n_spgz=sh_n_source, save_suffix='_spliited', debug=False)
                    st.write(f"Обрбатывается файл '{uploaded_file.name}' ...")
                    spgz_code_name, spgz_characteristics_content_loc_df = extract_spgz_df_lst_st(
                    fn=fn_proc_save,
                    sh_n_spgz=sh_n_source,
                    groupby_col='№п/п',
                    unique_test_cols=['Наименование СПГЗ', 'Единица измерения', 'ОКПД 2', 'Позиция КТРУ'],
                    # significant_cols = [
                    #     'Наименование характеристики', 'Единица измерения характеристики', 'Значение характеристики', 'Тип характеристики', 'Тип выбора значений характеристики заказчиком'],
                    significant_cols = [
                        'Наименование характеристики', 'Единица измерения характеристики', 'Значение характеристики', 
                        'Тип характеристики', 
                        'Тип выбора значений характеристики заказчиком',
                        'Условная операция',
                        'Код КТРУ характеристики', # 'Код характеристики КТРУ',

                        ],
                    )
                    if debug: 
                        st.write(spgz_code_name)
                        st.dataframe(spgz_characteristics_content_loc_df.head(2))
                    
                    
                    kpgz_head, chars_of_chars_df = create_kpgz_data(
                        spgz_characteristics_content_loc_df, debug = False)
                    if debug: 
                        st.write(kpgz_head)
                        st.dataframe(chars_of_chars_df.head(2))

                    fn_save = uploaded_file.name.split('.xlsx')[0] + '_upd.xlsx'
                    write_head_kpgz_sheet_st(
                            uploaded_file,
                            fn_save,
                            spgz_code_name,
                            kpgz_head,
                            chars_of_chars_df,
                            okpd2_df,
                            debug=False
                        )
                    fn_save_lst.append (fn_save)


        if len(fn_save_lst) > 0:
            fn_zip = "form_spgz.zip"
            with zipfile.ZipFile(fn_zip, "w") as zf:
                fn_save_lst = list(set(fn_save_lst))
                for fn_save in fn_save_lst:
                    zf.write(fn_save)
                    break
                zip_name_lst = zf.namelist()
            if len(fn_save_lst) > 1:
                with zipfile.ZipFile(fn_zip, "a") as zf:
                    fn_save_lst = list(set(fn_save_lst))
                    for fn_save in fn_save_lst:
                        zf.write(fn_save)
                    zip_name_lst = zf.namelist()
            st.write("Список обработанных файлов, сохраненных в zip-архиве:")
            st.write(zip_name_lst)
            
            st.write("Для скачивания zip-архива с обработанными файламинажмите на клавишу 'Download Zip'")
            with open(fn_zip, 'rb') as f:
                if st.download_button('Download Zip', f, mime='application/octet-stream', file_name=fn_zip):  # Defaults to 'application/octet-stream'
                    st.write('Работа программы завершена. Спасибо!')

    return

def main_03_v01(
    sh_n_source = 'СПГЗ',
    debug=False,
):

    okpd2_df = read_okpd_dict_fr_link()
    st.dataframe(okpd2_df.head(2)) #, use_container_width=True)

    uploaded_files = st.file_uploader(
        "Загрузите xlsx-файлы для обработки", accept_multiple_files=True
    )
    if uploaded_files:
        fn_lst = [fn.name for fn in  uploaded_files if fn.name.endswith('.xlsx')]
        fn_save_lst = []
        if len (fn_lst) == 0:
            st.write(f"В загруженных файлах не найдены .xlsx файлы")
            # st.write(f"Работа программы завершена")
            # st.write(f"Обновите страницу")
            # sys.exit(2)
        else:
            fn_save_lst = []
            for uploaded_file in uploaded_files:
                if uploaded_file.name.endswith('.xlsx'):
                    fn_proc_save = split_merged_cells_st(uploaded_file, sh_n_spgz=sh_n_source, save_suffix='_spliited', debug=False)
                    
                    spgz_code_name, spgz_characteristics_content_loc_df = extract_spgz_df_lst_st(
                    fn=fn_proc_save,
                    sh_n_spgz=sh_n_source,
                    groupby_col='№п/п',
                    unique_test_cols=['Наименование СПГЗ', 'Единица измерения', 'ОКПД 2', 'Позиция КТРУ'],
                    significant_cols = [
                        'Наименование характеристики', 'Единица измерения характеристики', 'Значение характеристики', 'Тип характеристики', 'Тип выбора значений характеристики заказчиком'],
                    )
                    if debug: 
                        st.write(spgz_code_name)
                        st.dataframe(spgz_characteristics_content_loc_df.head(2))
                    
                    
                    kpgz_head, chars_of_chars_df = create_kpgz_data(
                        spgz_characteristics_content_loc_df, debug = False)
                    if debug: 
                        st.write(kpgz_head)
                        st.dataframe(chars_of_chars_df.head(2))

                    fn_save = uploaded_file.name.split('.xlsx')[0] + '_upd.xlsx'
                    write_head_kpgz_sheet_st(
                            uploaded_file,
                            fn_save,
                            spgz_code_name,
                            kpgz_head,
                            chars_of_chars_df,
                            okpd2_df,
                            debug=False
                        )
                    fn_save_lst.append (fn_save)


        if len(fn_save_lst) > 0:
            fn_zip = "form_spgz.zip"
            with zipfile.ZipFile(fn_zip, "w") as zf:
                fn_save_lst = list(set(fn_save_lst))
                for fn_save in fn_save_lst:
                    zf.write(fn_save)
                    break
                zip_name_lst = zf.namelist()
            if len(fn_save_lst) > 1:
                with zipfile.ZipFile(fn_zip, "a") as zf:
                    fn_save_lst = list(set(fn_save_lst))
                    for fn_save in fn_save_lst:
                        zf.write(fn_save)
                    zip_name_lst = zf.namelist()
            st.write("Список обработанных файлов, сохраненных в zip-архиве:")
            st.write(zip_name_lst)
            
            st.write("Для скачивания zip-архива с обработанными файламинажмите на клавишу 'Download Zip'")
            with open(fn_zip, 'rb') as f:
                if st.download_button('Download Zip', f, mime='application/octet-stream', file_name=fn_zip):  # Defaults to 'application/octet-stream'
                    st.write('Работа программы завершена. Спасибо!')

    return

main_03(debug=True)