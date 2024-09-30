import pandas as pd
import numpy as np
import os, sys, glob
import humanize
import re
import xlrd

import json
import itertools
import requests
# from urllib.parse import urlencode
#from urllib.request import urlopen
#import requests, xmltodict
import time, datetime
import math
from pprint import pprint
import gc
from tqdm import tqdm
tqdm.pandas()
import pickle

import logging
import zipfile
import warnings
import argparse

import warnings
warnings.filterwarnings("ignore")

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import units
from openpyxl.styles import Border, Side, PatternFill, GradientFill, Alignment
from openpyxl import drawing

# from matplotlib.colors import ListedColormap, BoundaryNorm


class Logger():
    def __init__(self, name = 'Fuzzy Lookup',
                 strfmt = '[%(asctime)s] [%(levelname)s] > %(message)s', # strfmt = '[%(asctime)s] [%(name)s] [%(levelname)s] > %(message)s'
                 level = logging.INFO,
                 datefmt = '%H:%M:%S', # '%Y-%m-%d %H:%M:%S'
                #  datefmt = '%H:%M:%S %p %Z',

                 ):
        self.name = name
        self.strfmt = strfmt
        self.level = level
        self.datefmt = datefmt
        self.logger = logging.getLogger(name)
        self.logger.setLevel(self.level) #logging.INFO)
        self.offset = datetime.timezone(datetime.timedelta(hours=3))
        # create console handler and set level to debug
        self.ch = logging.StreamHandler()
        self.ch.setLevel(self.level)
        # create formatter
        self.strfmt = strfmt # '[%(asctime)s] [%(levelname)s] > %(message)s'
        self.datefmt = datefmt # '%H:%M:%S'
        # СЃРѕР·РґР°РµРј С„РѕСЂРјР°С‚С‚РµСЂ
        self.formatter = logging.Formatter(fmt=strfmt, datefmt=datefmt)
        self.formatter.converter = lambda *args: datetime.datetime.now(self.offset).timetuple()
        self.ch.setFormatter(self.formatter)
        # add ch to logger
        self.logger.addHandler(self.ch)
logger = Logger().logger
logger.propagate = False

if len(logger.handlers) > 1:
    for handler in logger.handlers:
        logger.removeHandler(handler)
    # del logger
    # logger = Logger().logger
    # logger.propagate = False


def unzip_file(path_source, fn_zip, work_path):
    logger.info('Unzip ' + fn_zip + ' start...')

    try:
        with zipfile.ZipFile(path_source + fn_zip, 'r') as zip_ref:
            fn_list = zip_ref.namelist()
            zip_ref.extractall(work_path)
        logger.info('Unzip ' + fn_zip + ' done!')
        return fn_list[0]
    except Exception as err:
        logger.error('Unzip error: ' + str(err))
        sys.exit(2)

def save_df_lst_to_excel_xlsxwriter(
    df_lst,
    data_processed_dir,
    fn_main,
    sh_n_lst,
    widths_lsts_list = None,
    indexes = None,
    ):

    offset = datetime.timezone(datetime.timedelta(hours=3))
    dt = datetime.datetime.now(offset)
    str_date = dt.strftime("%Y_%m_%d_%H%M")
    fn_save = fn_main + '_' + str_date + '.xlsx'

    if widths_lsts_list is None:
        widths_lsts_list = len(df_lst) * [[]]
    elif len(widths_lsts_list) < len(df_lst):
        widths_lsts_list.extend((len(df_lst) - len(widths_lsts_list)) * [[]])
    else:
        widths_lsts_list = [sub_lst for i_s, sub_lst in  enumerate(widths_lsts_list) if i_s < len(df_lst)]

    if indexes is None:
        indexes = len(df_lst) * [False]
    elif len(indexes) < len(df_lst):
        indexes.extend ( (len(df_lst)-len(indexes)) * [False])
    else:
        indexes = [item for i, item in  enumerate(indexes) if i < len(df_lst)]

    logger.info(f"Файл '{fn_save}' - старт записи в Excel...")
    with pd.ExcelWriter(os.path.join(data_processed_dir, fn_save), engine='xlsxwriter') as writer:
        workbook = writer.book
        format_float = workbook.add_format({"num_format": "# ### ##0.00"})
        format_int = workbook.add_format({"num_format": "# ### ##0"})
        header_format = workbook.add_format({'bold': True,"text_wrap": 1,"valign": "top", "align": "left",}) #'fg_color': '#D7E4BC','border': 1})

        for sh_n, data_df, idx, cols_width  in zip(sh_n_lst, df_lst, indexes, widths_lsts_list):
            data_df.to_excel(writer, sheet_name = sh_n, float_format="%.2f", index=idx) #
            worksheet = writer.sheets[sh_n]
            # print(cols_width)
            for i_w, w in enumerate(cols_width):
                worksheet.set_column(i_w, i_w, w, None)
            worksheet.autofilter(0, 0, data_df.shape[0], data_df.shape[1]-1)
            logger.info(f"Лист '{sh_n}' сохранен")
    logger.info(f"Обработанный файл '{fn_save}' сохранен в папке '{data_processed_dir}'")
    # !du -h "$data_processed_dir"/"$fn_save"
    file_size = get_humanize_filesize(data_processed_dir, fn_save)
    logger.info(f"Размер файла - {file_size}")

    return fn_save

def save_df_to_excel(df, path_to_save, fn_main, columns = None, b=0, e=None, index=False):
    offset = datetime.timezone(datetime.timedelta(hours=3))
    dt = datetime.datetime.now(offset)
    str_date = dt.strftime("%Y_%m_%d_%H%M")
    fn = fn_main + '_' + str_date + '.xlsx'
    logger.info(fn + ' save - start ...')
    if e is None or (e <0):
        e = df.shape[0]
    if columns is None:
        df[b:e].to_excel(os.path.join(path_to_save, fn), index = index)
    else:
        df[b:e].to_excel(os.path.join(path_to_save, fn), index = index, columns = columns)
    logger.info(fn + ' saved to ' + path_to_save)
    hfs = get_humanize_filesize(path_to_save, fn)
    logger.info("Size: " + str(hfs))
    return fn

def save_df_lst_to_excel(df_lst, sheet_names_lst, save_path, fn):
    # fn = model + '.xlsx'
    offset = datetime.timezone(datetime.timedelta(hours=3))
    dt = datetime.datetime.now(offset)
    str_date = dt.strftime("%Y_%m_%d_%H%M")
    fn_date = fn.replace('.xlsx','')  + '_' + str_date + '.xlsx'

    # with pd.ExcelWriter(os.path.join(path_tkbd_processed, fn_date )) as writer:
    with pd.ExcelWriter(os.path.join(save_path, fn_date )) as writer:

        for i, df in enumerate(df_lst):
            df.to_excel(writer, sheet_name = sheet_names_lst[i], index=False)
    return fn_date

def tar_file(data_source_dir, fn, data_processed_dir):
    fn_tar_gz = f'{fn}.tar.gz'
    logger.info(f"Упаковка файла '{fn}' - начало...")
    with tarfile.open(os.path.join(data_processed_dir, fn_tar_gz), 'w:gz') as tar:
        tar.add(os.path.join(data_source_dir, fn), arcname=fn)
    tar.close()
    file_size = get_humanize_filesize(data_processed_dir, fn_tar_gz)
    logger.info(f"Упаковка файла '{fn_tar_gz}' - завершено!")
    logger.info(f"Размер файла - {file_size}")

    return fn_tar_gz

def get_humanize_filesize(path, fn):
    human_file_size = None
    try:
        fn_full = os.path.join(path, fn)
    except Exception as err:
        print(err)
        return human_file_size
    if os.path.exists(fn_full):
        file_size = os.path.os.path.getsize(fn_full)
        human_file_size = humanize.naturalsize(file_size)
    return human_file_size

def restore_df_from_pickle(path_files, fn_pickle):

    if fn_pickle is None:
        logger.error('Restore pickle from ' + path_files + ' failed!')
        sys.exit(2)
    if os.path.exists(os.path.join(path_files, fn_pickle)):
        df = pd.read_pickle(os.path.join(path_files, fn_pickle))
        # logger.info('Restore ' + re.sub(path_files, '', fn_pickle_СЃ) + ' done!')
        logger.info('Restore ' + fn_pickle + ' done!')
        logger.info('Shape: ' + str(df.shape))
    else:
        # logger.error('Restore ' + re.sub(path_files, '', fn_pickle_СЃ) + ' from ' + path_files + ' failed!')
        logger.error('Restore ' + fn_pickle + ' from ' + path_files + ' failed!')
    return df


def insert_pd_col_after_col(
    df, move_col, trgt_col
):
    flag_miss_col = False
    if move_col not in df.columns:
        flag_miss_col=True
        print(f"Колонка '{move_col}' отсутсвует в DataFrame")
    if trgt_col not in df.columns:
        flag_miss_col=True
        print(f"Колонка '{trgt_col}' отсутсвует в DataFrame")
    if flag_miss_col:
        logger.error("Работа программы 'insert_pd_col_after_col' прекращена")
        sys.exit()

    df_columns = list(df.columns)
    # if trgt_col in df_columns and set(move_cols).issubset(df.columns):
    if trgt_col in df_columns and move_col in df_columns:
        trgt_col_idx = df_columns.index(trgt_col)
        move_col_idx = df_columns.index(move_col)
        if trgt_col_idx >= move_col_idx:
            part_cols_01 = df_columns[:move_col_idx]
            part_cols_02 = df_columns[move_col_idx + 1 : trgt_col_idx+1]
            part_cols_03 = [move_col]  + df_columns[trgt_col_idx + 1:]
            new_cols_order = part_cols_01 + part_cols_02 + part_cols_03
        else:
            part_cols_01 = df_columns[:trgt_col_idx+1]
            part_cols_02 = [move_col] + df_columns[trgt_col_idx + 1: move_col_idx]
            part_cols_03 = df_columns[move_col_idx + 1:]
            new_cols_order = part_cols_01 + part_cols_02 + part_cols_03
        return df[new_cols_order]
    else:
        print(f"Указанные колонки '{move_col}', '{trgt_col}' отсутствуют в DataFrame")
        return df

last_sym_to_cut = [ ',', '.', ':', ';', '_', '\xa0','\\']
import unicodedata
def clean_str(s, to_lower=False):
    """
    v 01.03 01.07.2024
        upd
        replace ('й', 'й') - замена двухсимволной 'й' на односимвольную
        replace (chr(160), chr(32)) - замена неразрывного пробла на обычный
    """
    if type(s)==str:
        s = unicodedata.normalize("NFKD", s)
        s = re.sub(r" +", ' ', s.strip())
        if (len(s) > 0) and s[-1] in last_sym_to_cut:
            return clean_str(s[:-1], to_lower)
        else:
            # s_clean = s.strip().replace('й', 'й').replace('Й', 'Й').replace(chr(160), ' ') # 'й' из двух символов непереносимые пробелы
            s_clean = s.strip().replace('й', 'й').replace('Й', 'Й').replace(chr(160), ' ').replace('Ё', 'Е').replace('ё', 'е').replace('ё', 'е')

            if to_lower:
                s_clean = s_clean.lower()
            return s_clean
    return s

def split_column_to_rows(
    df,
    to_split_col = 'ИНП',
    splitted_col = 'ИНП split',
    sep = ', ',
    rename_col = True,
    # new_df = True,
    debug = False
    ):

    flag_miss_col = False
    if to_split_col not in df.columns:
        flag_miss_col=True
        print(f"Колонка '{to_split_col}' отсутсвует в DataFrame")
    # if splitted_col not in df.columns:
    #     flag_miss_col=True
    #     print(f"Колонка '{splitted_col}' отсутсвует в DataFrame")
    if flag_miss_col:
        logger.error("Работа программы 'split_column_to_rows' прекращена")
        sys.exit()

    df_02_lst = []
    for i_row, row in tqdm(df.iterrows(), total = df.shape[0]):
        val_to_split = row[to_split_col]
        dict_row = dict(row)
        if type(val_to_split)==str:

            for v in val_to_split.split(sep):
                dict_row_cpy = dict_row.copy()
                dict_row_cpy.update(
                    {splitted_col: clean_str(v)}
                )
                df_02_lst.append(dict_row_cpy)
        else:
            dict_row.update(
                {splitted_col: val_to_split}
            )
            df_02_lst.append(dict_row)
        if debug and (i_row > 2): break
    df_02_lst[:5]
    if len(df_02_lst) > 0:
        df_02 = pd.DataFrame(df_02_lst)
    else:
        df_02= pd.DataFrame(df_02_lst)
        # ???
    if debug:
        print(df_02.shape)
    if rename_col:
        df_02.drop(columns=[to_split_col], inplace=True)
        df_02.rename(columns={splitted_col: to_split_col}, inplace=True)
    return df_02

def get_cols_width_exists(dir, fn, sh_n):
    """
    только для xlsx (xlsb Не поддерживает)
    """
    wb = load_workbook(os.path.join(dir,fn))
    ws = wb.get_sheet_by_name(sh_n)
    cols_width_exists = []
    # ws.sheet_state, ws.max_row, ws.max_column
    for ic in range(ws.max_column):
        cell = ws.cell(row=1, column=ic+1)
        cols_width_exists.append(ws.column_dimensions[cell.column_letter].width)
    return cols_width_exists

# import py7zr
# def un_7zip(data_source_dir, fn_zip):
#     # if zipfile.is_zipfile(os.path.join(data_source_dir, fn_zip)):
#         # with zipfile.open(os.path.join(data_source_dir, fn_zip), 'r') as tar:
#     # else:
#     #     print(f"'{fn_tar_gz}' is not zip file")
#     with py7zr.SevenZipFile(os.path.join(data_source_dir, fn_zip), 'r') as z:
#         print(z.list())
#         z.extractall(path=data_source_dir)
